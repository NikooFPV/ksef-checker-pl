import tkinter as tk
from tkinter import ttk, filedialog
import threading, os, sys, base64, io, webbrowser
import pandas as pd

VERSION     = "2.1.2"
GITHUB_REPO = "NikooFPV/ksef-checker-pl"
GITHUB_API  = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
GITHUB_URL  = f"https://github.com/{GITHUB_REPO}/releases/latest"

def _ver_tuple(v):
    try: return tuple(int(x) for x in v.lstrip("v").split("."))
    except: return (0,)

import config as cfg_module
from config import ALL_CHECKS

# ── logo (base64 PNG — brak zewnętrznych plików) ─────────────────────────────
_LOGO_48 = "R0lGODdhMAAwAIIAAP////Dw8L6+vrS0tPuSPCIiIhoaGgAAACwAAAAAMAAwAEAI/wANCBxIsKDBgwgTGizAsGGBABAjSpw4wKHFixYJYnw4saPHiBsxDtz4sWQAAihTpgxZYCRGkxFVogSQEgBNmQQyumTJkyHOlRsP9hxKdOFLmBEHKFVKtOFOixIFSJ1KlarEpk8dIoX4syvKi1kbbj2J8+ZPsAWbqg2qUODaom3fynWY9ujWpXLDMhwbYKlfrG7t8r0K1wDJiFUTTyU8VC9HiIqpesUpMvDFqJEnE7Bp9itdy1CRluXcWaVOw3N9agZ6GnRqtW1Rv64cu7bt20Jn81Soe63R3rAd82Wa17XYsX99G9+L3G9FwLJDD2bcU/h06ruXPw4QOTF2ltavQ1GEfhhyd6tkWYcMv3V1TtrluZ937/l0/PmTS6ONrlW0V86UfcbfcTDJBKBNXdknmEmjlaYeQ441Rd97Cr7mHlvaySVTdhoBV1huHu6H24gkBgQAOw=="
_LOGO_256 = "R0lGODdhAAEAAYIAAP////Dw8L6+vrS0tPuSPCIiIhoaGhISEiwAAAAAAAEAAUAI/wANCBxIsKDBgwgTKlzIsKHDhxAjSpxIsaLFixgzatzIsaPHjyA3FhhJsqTJkyhTqlzJsqXLlzBjypwZsqbBmThz6tzJs6dPnjYl/hxKtKjRo0hTBiWYtKnTp1CjKv0otWSAq1izat3KtavXr2C7Dhg7tqrZniLPjgzLtq3bt1nJyp07QK3dlhjvFoDLt69fuoAD69V7Ua/fw4jBBl7MuOzgp4XvJp5M+Wrjy5jrPh4a2W7lz6BDB9jM2aJh0ahTwyX9s7NavwJiy55Nu7bt27hz69aNmLVP12dVCx/u1Tda05KJK1duHChyz3x3S59OvbqAAASya9/Ovbv37+C73/8Fbna5eazh06tfz359U/JVo1ufT19A+/v48+s/ULri6fOf6SfggNsBYKCBBG7Hn07wSSVffRDSluCE7B1o4YUXJrigTA1GBWBiFIbYHYYkWiiighw+99qHh51IQIkmfgcjjC5+d8CGLnUIFYuIuTjjjzPWuN6NKfqXHI8tCqkdkBkqyR6RNOnYnG9OVjnhjTgWmdeUXKZk5Zc2YnmcR12WuRKYIWKZ5ZhLmemmS2oKqaaakC310Jt45qmSnTbp6edZfAYq6KCEFmrooYgmquiijGr056MrNRoopJQ+JilClWaqp6Gadqppn+UhCZpmniLF0ZGiUrZYqURtCV2qqmb/5hirURr5KqyJycoYrSxJ6RSuseoqK68j+doUsMh+xaqxSSXr7FbLqhjcs9SOViqzSD0Y4bbcXtdXtLauWK2z4FL0n1vdphvhV0JOie1R44KGJoVSvWsUbOrma9+8867J5kTnxusVvwSH56+WAKMqcFYFD8ikhjnZW1Rf+lrX8H1MxkghlDBJTNTCDCvZZIUZI+gkxzlKGyrI2IVY8oHZvWwymijjpXJ8LKMnpMwkXkyAmDF5PFTOW335o8/aAV1ruMQ+hXS/SkecVtNVPZ1m1DuBSvVdB1idNJ1Hcbq1m3OWDfZ4l9409toIpw0S22W6LffcdNdt991476333nz3/+3334AHLijcnQruEOGIlwR44ozbPHfjkC+daOSUS11o5ZhnzWfmnP9LZuegO9eRuERzRVboJ51KeulaAYZ6sRkpzDpWjYXu6uqzW5YZ50L/lPtWws7aeO8+/d568HIxTnxPxseFvOtsL89T87Q/vyvV0u9Eve7WX8Zr9jptv1y5Cd8qvmrkCyX7+aKlH1HA7Ld/7c0Oxo/+/EyvbL/8noKfk7YVC+Bt/OI+iMCPLQJMYG0OU8A7rW9/lGng4R4IwcRIsCEHRKACFeiX9uTJfzgB4AYTGDKCbQaEM6kgV7z2HUDRz0MqZCF+oIJCmdhPhiEy1Qt3BJcRQgiH7WrVDv+fQjEfUgeIX2rNEH/1OySmBwAiOljHlnis2TkRPDSCmOXMRUGwGBE3VxzRywRUs5fUMCYi/GJswligMV5Jcuoz38LCyLMTlbFXVGxWzqwExfzUUU5SRMkZYUK0nZVIPTL70h33lMds7dFlJRNjJGkWyNQ1El6FFBHP3Ai1SppkkC8pnSE32bOGLZKR+cOZKEVGSp9hLWWprB/r+Jixp70Sllx8HUvQNDJbntJxsdSlSdgISE9GKnbCdAkxr/RLM04tmS1ZJn5uiROqQBMmXZMmd+bUn7ddUybZdCI3izKpb+LkRqYcZ9jEZs6elK09ZqOmDhfVTuXRrZ7EWhw+/WT/OAzu0y79DKhAB0rQghr0oAhNqEIXytCGOvShEI2oRCdK0Ypa9KIYzahGVffPOlG0oydUKEjjJtCRftCfJv3T3lJauLqx9ILsfOn3GCXTrSmqptG7HE4JN6idJq6cPkVcm4IKuZoQlXLePGrkrKnK5gnvdZ/Tn/GSJ0yOSvV3dKmqo3CXO8Ggbqtcnd31QIfMsLLOe5273bSotzveXfJe29NV5kDpEvEFr3J0bYldkbfUt04srtYbnl8/Btju/XSwQyts90i1tryyZK+LPV1jEes7xUb2qTMNJg/ZetmsNs2xK4FsZ+eST8oWz7KjpSr+cmlWFRJwteVrrWv5AtOF/2Rwtt+CbRxli1u31FYht+3tW36bkOAKty3ExVQXj8uW5B7EuMwFi3PVJsfo+la374OudauDXQNqd7tcmW5BoAteaHXXgdUNixoDyMDzTjC9X1mvAF/bP9Myr4fyrRh9W6pZIhYxv9zqjXv9Cd/ydkW8TPmugbGC4IGQd8FXabBAHgxhCRuAwgu2MIaxAmB1tYVeXQKtStLYYQitMImsEXFKSFxi+lzlaYNRMUrw1eL6tEyGALXv9HpLTLPI+CSz1aZ2ovJjk8RQyNzxKGvXyj4kpycpRbZK/JysHmPCMbsF9kqNwUhl9hglyiRhcYu7nB8h9peJb9lybci8HyWeuf+KYpYvm8n4Gx1rr4lzJpDnsMzbcTkZZmnS3Jv1OEskl1KL1bRz+KyozaMRqJmoXPJVQdboSbY50YN2JH47XGlLX7ptfGZyW2q8zD8OCNKWzDQm01ziUnMS0VNUNVzjrEA2JjKKViYJmNfyyCve2o6g9m6WneXrV9sx1wXY9V56PUr8mLpGqNa1ov/H7BMd+onGhjaylZ1Ja2cRi78+2banHcJqJ6iW3gl3laINO1n/lWWQfLW6rcTuZJM7heYmUB3nTe9xu5uw+dY3KR3NL3kK8t427LYmB37IgtXb3v9OrMIXzvCZEczgB494ZSfu7YG7EtntlnRTAx7vbBf84RDQF7ksV8lKdH8c5CGPrahJTvEg+XKLMjcnmBpuNYxHOuc6z/OpUS5thBNO6NP0+c9VHnSkV5noiivrP52eHqUfU+odpXp3zsagZ4JU6z+zOh6t+nWni33sZB9pOKmszj1jPaVO5nqdkyrTtYtT7nM36k7t3nO85z0oROX7xf3u5s0dFUtgMptTYnp4xDOzbESenFJPEs/Kx9OF9Jx8pdym+U25tPMkzRvofRO40aMtoKaPfEJT302Jst6ZG4297GdP+9rb/va4z73ud8/73u8tIAA7"

def _logo_photoimage(b64: str):
    """Base64 GIF → tk.PhotoImage (GIF obsługiwany natywnie przez Tk)."""
    return tk.PhotoImage(data=b64)

# ── mapowanie id sprawdzenia → kategoria (do grupowania na liście wyników)
_CAT_MAP  = {cid: cat for cid, _, cat, _ in ALL_CHECKS}
_CAT_ORDER = list(dict.fromkeys(cat for _, _, cat, _ in ALL_CHECKS))

# ── palety motywów ───────────────────────────────────────────────────────────
THEMES = {
    "dark": {
        "BG":"#18181b","BG2":"#1f1f23","BG3":"#272730","BG4":"#313139",
        "BORDER":"#2e2e38","BG_HOVER":"#25252f",
        "TXT":"#fafafa","TXT2":"#a1a1aa","TXT3":"#52525b",
        "CARD_ERR":"#231010","CARD_WARN":"#251d08",
        "MARKED":"#1a2e1e","MARKED_FG":"#86efac",
        "THEME_ICON":"☀️","THEME_LABEL":"Jasny",
    },
    "light": {
        "BG":"#f6f8fa","BG2":"#ffffff","BG3":"#f0f2f4","BG4":"#e6e8eb",
        "BORDER":"#d0d7de","BG_HOVER":"#eaeef2",
        "TXT":"#24292f","TXT2":"#57606a","TXT3":"#8c959f",
        "CARD_ERR":"#fff1f2","CARD_WARN":"#fffbeb",
        "MARKED":"#e6f4ea","MARKED_FG":"#1a6e3a",
        "THEME_ICON":"🌙","THEME_LABEL":"Ciemny",
    },
}

# ── aktywna paleta (aktualizowana przez apply_theme) ─────────────────────────
BG    = "#18181b"; BG2 = "#1f1f23"; BG3 = "#272730"; BG4 = "#313139"
ACCENT= "#f97316"; A2  = "#22c55e"; WARN= "#eab308"; ERR = "#ef4444"
OK    = "#22c55e"; TXT = "#fafafa"; TXT2= "#a1a1aa"; TXT3= "#52525b"
BORDER= "#2e2e38"; BG_HOVER = "#25252f"
CARD_ERR = "#231010"; CARD_WARN = "#251d08"
MARKED = "#1a2e1e"; MARKED_FG = "#86efac"

def apply_theme(name: str):
    """Zaktualizuj globalne zmienne kolorów wg wybranego motywu."""
    global BG,BG2,BG3,BG4,BORDER,BG_HOVER,TXT,TXT2,TXT3,CARD_ERR,CARD_WARN,MARKED,MARKED_FG
    t = THEMES.get(name, THEMES["dark"])
    BG=t["BG"]; BG2=t["BG2"]; BG3=t["BG3"]; BG4=t["BG4"]
    BORDER=t["BORDER"]; BG_HOVER=t["BG_HOVER"]
    TXT=t["TXT"]; TXT2=t["TXT2"]; TXT3=t["TXT3"]
    CARD_ERR=t["CARD_ERR"]; CARD_WARN=t["CARD_WARN"]
    MARKED=t["MARKED"]; MARKED_FG=t["MARKED_FG"]

# ── fonts ─────────────────────────────────────────────────────────────────────
# Segoe UI Variable = Windows 11 modern, falls back to Segoe UI on Win10
_SYS  = "Segoe UI Variable"
FMONO = "Cascadia Code"        # monospace — fallback do Consolas
FB    = (_SYS, 11);  FS = (_SYS, 10);  FSM = (_SYS, 9)
FBIG  = (_SYS, 20, "bold");    FMED = (_SYS, 12, "bold")

MONTHS_PL = ["Styczeń","Luty","Marzec","Kwiecień","Maj","Czerwiec",
             "Lipiec","Sierpień","Wrzesień","Październik","Listopad","Grudzień"]

# ── database ──────────────────────────────────────────────────────────────────
def open_connection(mdb_path):
    try: import pyodbc
    except ImportError:
        raise RuntimeError("Brak pyodbc.\n\nUruchom:\n    python -m pip install pyodbc")
    for drv in [r"Microsoft Access Driver (*.mdb, *.accdb)",
                r"Microsoft Access Driver (*.mdb)"]:
        try:
            return pyodbc.connect(
                f"DRIVER={{{drv}}};DBQ={mdb_path};ExtendedAnsiSQL=1;",
                autocommit=True)
        except Exception: continue
    raise RuntimeError(
        "Brak sterownika Microsoft Access.\n\n"
        "Pobierz AccessDatabaseEngine_X64.exe:\n"
        "https://www.microsoft.com/en-us/download/details.aspx?id=54920")

def read_table(conn, name):
    try:
        df = pd.read_sql(f"SELECT * FROM [{name}]", conn)
        return df.astype(str).replace({"None":"","nan":"","<NA>":""})
    except Exception: return None

def list_tables(conn):
    return [t.table_name for t in conn.cursor().tables(tableType="TABLE")]

# ── analysis wrapper ──────────────────────────────────────────────────────────
def analyze_mdb(mdb_path, month=None, year=None, cfg=None, progress_cb=None, quarter=None):
    def prog(msg):
        if progress_cb: progress_cb(msg)
    prog("Łączę z bazą…")
    conn = open_connection(mdb_path)
    try:
        prog("Odczytuję strukturę…")
        tables = list_tables(conn)
        for req in ["_KSEF_DOCUMENT","KSIEGA","VATZAKUPY"]:
            if req not in tables:
                return {"status":"error","checks":[{"kind":"error",
                    "title":f"Nieobsługiwany format bazy — brak tabeli {req}.",
                    "detail":"","rows":[],"explanation":""}],"summary":{},"period":"?"}
        prog("Wczytuję dane…")
        ksef   = read_table(conn, "_KSEF_DOCUMENT")
        ksiega = read_table(conn, "KSIEGA")
        vat    = read_table(conn, "VATZAKUPY")
        vatsp  = read_table(conn, "VATSPRZEDAZ") if "VATSPRZEDAZ" in tables else None
    finally:
        conn.close()  # zawsze zamknij — nawet przy błędzie (usuwa pliki .dat)
    if ksef is None or ksiega is None or vat is None:
        return {"status":"error","checks":[{"kind":"error","title":"Błąd odczytu tabel.",
            "detail":"","rows":[],"explanation":""}],"summary":{},"period":"?"}
    from checks import run_all
    return run_all(ksef, ksiega, vat, vatsp, month, year,
                   cfg=cfg, prog_cb=progress_cb, quarter=quarter)

def detect_latest_period(mdb_path):
    """Najnowszy (miesiąc, rok) z POBRANYCH dokumentów KSeF.

    Domyślny okres po otwarciu pliku — sensowniejszy niż bieżący miesiąc
    kalendarzowy, który dla zamkniętego okresu bywa pusty (faktury jeszcze
    nie pobrane z KSeF → DOWNLOADED=False → odfiltrowane). Zwraca None gdy
    brak pobranych dokumentów."""
    try:
        conn = open_connection(mdb_path)
    except Exception:
        return None
    try:
        ksef = read_table(conn, "_KSEF_DOCUMENT")
    except Exception:
        ksef = None
    finally:
        conn.close()
    if ksef is None or ksef.empty:
        return None
    if "DOWNLOADED" in ksef.columns:
        dl = ksef["DOWNLOADED"].astype(str).str.strip().str.upper()
        ksef = ksef[dl.isin(["1", "TRUE", "YES", "-1"])]
    if ksef.empty:
        return None
    iss = pd.to_datetime(ksef.get("ISSUE_DATE"), errors="coerce")
    inv = pd.to_datetime(ksef.get("INVOICING_DATE"), errors="coerce")
    # data wystawienia decyduje o okresie — INVOICING_DATE (wysyłka do KSeF) tylko fallback
    eff = iss.fillna(inv).dropna()
    if eff.empty:
        return None
    mx = eff.max()
    return (int(mx.month), int(mx.year))

# ── settings window ───────────────────────────────────────────────────────────
class SettingsWindow(tk.Toplevel):
    def __init__(self, parent, cfg, on_save, app_ref=None):
        super().__init__(parent)
        self.title("Ustawienia — KSeF Checker")
        self.geometry("680x580")
        self.configure(bg=BG)
        self.resizable(True, True)
        self._cfg = dict(cfg)
        self._on_save = on_save
        self._app_ref = app_ref
        self._vars = {}
        self._build()
        self.lift(); self.focus_force()

    def _build(self):
        # header
        hdr = tk.Frame(self, bg=BG2, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="⚙  Ustawienia", font=FMED, bg=BG2, fg=ACCENT, padx=16).pack(side="left")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12, pady=8)

        style = ttk.Style(self)
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", background=BG3, foreground=TXT2,
                        font=FS, padding=[12,4])
        style.map("TNotebook.Tab", background=[("selected",BG2)],
                  foreground=[("selected",ACCENT)])

        # ── Tab 1: Sprawdzenia ────────────────────────────────────────────
        tab_checks = tk.Frame(nb, bg=BG)
        nb.add(tab_checks, text="  Sprawdzenia  ")

        tk.Label(tab_checks,
                 text="Wybierz które sprawdzenia mają być wykonywane:",
                 font=FS, bg=BG, fg=TXT2, anchor="w").pack(fill="x", padx=12, pady=(8,4))

        # scrollable list
        cv = tk.Canvas(tab_checks, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(tab_checks, orient="vertical", command=cv.yview)
        cv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        cv.pack(side="left", fill="both", expand=True, padx=(12,0))
        inner = tk.Frame(cv, bg=BG)
        win_id = cv.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.bind("<Configure>", lambda e: cv.itemconfig(win_id, width=e.width))
        cv.bind("<MouseWheel>", lambda e: cv.yview_scroll(-1*(e.delta//120),"units"))

        enabled = self._cfg.get("enabled_checks", {})
        current_cat = [None]

        for cid, default_on, cat, label in ALL_CHECKS:
            if cat != current_cat[0]:
                current_cat[0] = cat
                cat_frame = tk.Frame(inner, bg=BG3)
                cat_frame.pack(fill="x", pady=(8,2), padx=4)
                tk.Label(cat_frame, text=f"  {cat}", font=("Consolas",9,"bold"),
                         bg=BG3, fg=ACCENT, pady=3).pack(anchor="w")

            var = tk.BooleanVar(value=enabled.get(cid, default_on))
            self._vars[cid] = var
            row = tk.Frame(inner, bg=BG2)
            row.pack(fill="x", padx=4, pady=1)
            tk.Checkbutton(row, variable=var, bg=BG2, fg=TXT,
                           activebackground=BG2, selectcolor=BG3,
                           font=FS).pack(side="left")
            tk.Label(row, text=label, font=FS, bg=BG2, fg=TXT,
                     anchor="w").pack(side="left", padx=4)

        # bulk buttons
        btn_row = tk.Frame(tab_checks, bg=BG, pady=4)
        btn_row.pack(fill="x", padx=12)
        for lbl, cmd in [
            ("Zaznacz wszystkie", lambda: [v.set(True)  for v in self._vars.values()]),
            ("Odznacz wszystkie", lambda: [v.set(False) for v in self._vars.values()]),
            ("Przywróć domyślne", self._restore_defaults),
        ]:
            tk.Button(btn_row, text=lbl, font=FSM,
                      bg=BG4, fg=TXT,
                      activebackground=BG3, activeforeground=TXT,
                      relief="flat", padx=10, pady=4).pack(side="left", padx=2)
        # re-bind commands after creation (lambda issue workaround)
        btns = btn_row.winfo_children()
        btns[0].config(command=lambda: [v.set(True)  for v in self._vars.values()])
        btns[1].config(command=lambda: [v.set(False) for v in self._vars.values()])
        btns[2].config(command=self._restore_defaults)

        # ── Tab 2: Parametry ──────────────────────────────────────────────
        tab_params = tk.Frame(nb, bg=BG)
        nb.add(tab_params, text="  Parametry  ")

        params = [
            ("amount_tolerance",  "Tolerancja różnic kwotowych (zł):",
             "Różnice mniejsze niż ta wartość są ignorowane. Domyślnie: 0.05"),
            ("ksef_no_threshold", "Próg ostrzeżenia 'brak KSeF' (%):",
             "Jeśli więcej % zakupów nie ma numeru KSeF, wyświetl ostrzeżenie. Domyślnie: 15"),
            ("cutoff_days",       "Dni bufora dla 'cały zakres' (dni):",
             "Ile ostatnich dni ignorować jako 'za świeże do zaksięgowania'. Domyślnie: 3"),
        ]
        self._param_vars = {}
        for key, label, tip in params:
            f = tk.Frame(tab_params, bg=BG)
            f.pack(fill="x", padx=16, pady=8)
            tk.Label(f, text=label, font=FS, bg=BG, fg=TXT,
                     anchor="w", width=42).pack(side="left")
            var = tk.StringVar(value=str(self._cfg.get(key,
                cfg_module.DEFAULTS.get(key,""))))
            self._param_vars[key] = var
            tk.Entry(f, textvariable=var, font=FS, bg=BG3, fg=TXT,
                     insertbackground=TXT, relief="flat", width=10).pack(side="left")
            tk.Label(f, text=f"  ← {tip}", font=FSM,
                     bg=BG, fg=TXT3, anchor="w").pack(side="left")

        # ── bottom save/cancel + aktualizacje ────────────────────────────
        bot = tk.Frame(self, bg=BG2, pady=8)
        bot.pack(fill="x", side="bottom")
        tk.Button(bot, text="  Zapisz  ", font=FMED,
                  bg=ACCENT, fg="#18181b",
                  activebackground="#ea6d05", activeforeground="#18181b",
                  relief="flat", padx=14, pady=7,
                  command=self._save).pack(side="right", padx=12)
        tk.Button(bot, text="  Anuluj  ", font=FB,
                  bg=BG3, fg=TXT,
                  activebackground=BG4, activeforeground=TXT,
                  relief="flat", padx=12, pady=7,
                  command=self.destroy).pack(side="right", padx=4)

        # przycisk sprawdzenia aktualizacji (lewa strona)
        self._upd_btn = tk.Button(bot, text="🔔  Sprawdź aktualizacje", font=FSM,
                  bg=BG3, fg=TXT,
                  activebackground=BG4, activeforeground=TXT,
                  relief="flat", padx=12, pady=7,
                  command=self._check_updates_now)
        self._upd_btn.pack(side="left", padx=12)
        self._upd_lbl = tk.Label(bot, text="", font=FSM, bg=BG2, fg=TXT2)
        self._upd_lbl.pack(side="left")

    def _check_updates_now(self):
        self._upd_btn.config(state="disabled")
        self._upd_lbl.config(text="Sprawdzam…", fg=TXT2)
        import threading
        def worker():
            import urllib.request, json
            try:
                req = urllib.request.Request(GITHUB_API,
                                             headers={"User-Agent": "KSeF-Checker"})
                with urllib.request.urlopen(req, timeout=6) as r:
                    data = json.loads(r.read())
                tag    = data.get("tag_name", "")
                dl_url = data.get("html_url", GITHUB_URL)
                if _ver_tuple(tag) > _ver_tuple(VERSION):
                    # szukaj URL instalatora
                    inst_url = None
                    for asset in data.get("assets", []):
                        if "installer" in asset["name"].lower() and "access" not in asset["name"].lower():
                            inst_url = asset["browser_download_url"]
                            break
                    if not inst_url:
                        inst_url = dl_url
                    def _do_update(t=tag, u=inst_url):
                        self.destroy()
                        if self._app_ref:
                            self._app_ref._start_update(t, u)
                    self.after(0, lambda: (
                        self._upd_lbl.config(
                            text=f"Dostępna wersja {tag}!", fg=WARN),
                        self._upd_btn.config(
                            text="  Aktualizuj  ", state="normal",
                            bg=ACCENT, fg="#18181b",
                            command=_do_update)
                    ))
                else:
                    self.after(0, lambda: (
                        self._upd_lbl.config(
                            text=f"Masz najnowszą wersję ({VERSION}) ✓", fg=OK),
                        self._upd_btn.config(state="normal")
                    ))
                # wyzeruj datę żeby przy starcie też sprawdzało
                self._cfg["last_update_check"] = ""
                cfg_module.save(self._cfg)
            except Exception:
                self.after(0, lambda: (
                    self._upd_lbl.config(text=f"Błąd: {type(e).__name__} — {str(e)[:60]}", fg=ERR),
                    self._upd_btn.config(state="normal")
                ))
        threading.Thread(target=worker, daemon=True).start()

    def _restore_defaults(self):
        defaults = cfg_module.DEFAULTS["enabled_checks"]
        for cid, var in self._vars.items():
            var.set(defaults.get(cid, True))

    def _save(self):
        self._cfg["enabled_checks"] = {cid: var.get() for cid, var in self._vars.items()}
        for key, var in self._param_vars.items():
            try:
                self._cfg[key] = float(var.get())
            except ValueError: pass
        cfg_module.save(self._cfg)
        self._on_save(self._cfg)
        self.destroy()

# ── batch processing tab ─────────────────────────────────────────────────────
class BatchTab(tk.Frame):
    """Zakładka do sprawdzania wielu baz MDB na raz."""

    def __init__(self, parent, cfg_getter, on_open_single=None, **kw):
        super().__init__(parent, bg=BG, **kw)
        self._cfg_getter = cfg_getter
        self._on_open_single = on_open_single  # callback(path, month, year)
        self._files = []      # list of (path, month, year)
        self._running = False
        self._results = []    # list of (path, res)
        self._build()

    def _build(self):
        # ── toolbar ───────────────────────────────────────────────────────
        ctrl = tk.Frame(self, bg=BG2)
        ctrl.pack(fill="x")
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        ctrl_inner = tk.Frame(ctrl, bg=BG2)
        ctrl_inner.pack(fill="x", padx=16, pady=10)

        tk.Button(ctrl_inner, text="＋  Dodaj bazy MDB", font=FMED,
                  bg=ACCENT, fg="#18181b",
                  activebackground="#ea6d05", activeforeground="#18181b",
                  relief="flat", padx=14, pady=7, cursor="hand2",
                  command=self._add_files).pack(side="left")

        tk.Button(ctrl_inner, text="✕  Wyczyść", font=FB,
                  bg=BG3, fg=TXT2, activebackground=BG4, activeforeground=TXT,
                  relief="flat", padx=12, pady=7, cursor="hand2",
                  command=self._clear_list).pack(side="left", padx=(6,0))

        tk.Frame(ctrl_inner, bg=BORDER, width=1).pack(side="left", fill="y", padx=14)

        # segmented period control — identyczny jak w single tab
        self._period_var = tk.StringVar(value="month")
        seg = tk.Frame(ctrl_inner, bg=BG3)
        seg.pack(side="left")

        def _set_period(val):
            self._period_var.set(val)
            is_m = val == "month"
            _b_all.config(bg=ACCENT if not is_m else BG3,
                          fg="#18181b" if not is_m else TXT2)
            _b_mon.config(bg=ACCENT if is_m else BG3,
                          fg="#18181b" if is_m else TXT2)

        _b_all = tk.Button(seg, text="Cały", font=FS,
            bg=BG3, fg=TXT2, activebackground=BG4,
            relief="flat", padx=14, pady=7, cursor="hand2",
            command=lambda: _set_period("all"))
        _b_all.pack(side="left")
        tk.Frame(seg, bg=BORDER, width=1).pack(side="left", fill="y")
        _b_mon = tk.Button(seg, text="Miesiąc", font=FS,
            bg=ACCENT, fg="#18181b", activebackground="#ea6d05",
            relief="flat", padx=14, pady=7, cursor="hand2",
            command=lambda: _set_period("month"))
        _b_mon.pack(side="left")

        import datetime; now = datetime.datetime.now()
        self._month_var = tk.StringVar(value=MONTHS_PL[now.month-1])
        self._year_var  = tk.StringVar(value=str(now.year))
        ttk.Combobox(ctrl_inner, textvariable=self._month_var,
                     values=MONTHS_PL, state="readonly",
                     width=11, font=FS).pack(side="left", padx=(8,4))
        tk.Spinbox(ctrl_inner, from_=2020, to=2035,
                   textvariable=self._year_var, width=5, font=FS,
                   bg=BG3, fg=TXT, buttonbackground=BG4,
                   insertbackground=TXT, relief="flat").pack(side="left")

        tk.Frame(ctrl_inner, bg=BORDER, width=1).pack(side="left", fill="y", padx=14)

        self._run_btn = tk.Button(ctrl_inner, text="▶  Sprawdź wszystkie", font=FMED,
                                   bg=ACCENT, fg="#18181b",
                                   activebackground="#ea6d05", activeforeground="#18181b",
                                   relief="flat", padx=16, pady=7, cursor="hand2",
                                   state="disabled", command=self._run_all)
        self._run_btn.pack(side="left", padx=(0,6))

        self._export_btn = tk.Button(ctrl_inner, text="↓  Excel zbiorczy", font=FB,
                                      bg=BG3, fg=A2,
                                      activebackground=BG4, activeforeground=A2,
                                      relief="flat", padx=14, pady=7, cursor="hand2",
                                      state="disabled", command=self._export_batch)
        self._export_btn.pack(side="left")

        self._status = tk.Label(ctrl_inner, text="", font=FS, bg=BG2, fg=TXT2)
        self._status.pack(side="left", padx=12)

        style = ttk.Style()
        style.configure("B.Horizontal.TProgressbar",
                        troughcolor=BG3, background=ACCENT,
                        bordercolor=BG3, lightcolor=ACCENT, darkcolor=ACCENT)
        self._prog = ttk.Progressbar(ctrl_inner, style="B.Horizontal.TProgressbar",
                                      mode="determinate", length=160)

        # ── split: left=file list, right=results ──────────────────────────
        split = tk.Frame(self, bg=BG)
        split.pack(fill="both", expand=True)

        # LEFT – file queue
        left = tk.Frame(split, bg=BG2, width=300)
        left.pack(side="left", fill="y"); left.pack_propagate(False)
        tk.Frame(split, bg=BORDER, width=1).pack(side="left", fill="y")

        # nagłówek kolejki
        q_hdr = tk.Frame(left, bg=BG2)
        q_hdr.pack(fill="x", padx=14, pady=(12,6))
        tk.Label(q_hdr, text="KOLEJKA PLIKÓW", font=(_SYS,8,"bold"),
                 bg=BG2, fg=TXT3).pack(side="left")

        lsb = ttk.Scrollbar(left, orient="vertical")
        lsb.pack(side="right", fill="y")
        self._file_list = tk.Listbox(left, bg=BG2, fg=TXT, font=FS,
                                      selectbackground=BG3, selectforeground=ACCENT,
                                      borderwidth=0, highlightthickness=0,
                                      yscrollcommand=lsb.set, activestyle="none")
        self._file_list.pack(side="left", fill="both", expand=True, padx=(14,0))
        lsb.config(command=self._file_list.yview)

        self._file_list.bind("<Button-3>",
            lambda e: self._remove_file(self._file_list.nearest(e.y)))
        tk.Label(left, text="Prawy klik = usuń", font=(_SYS,8),
                 bg=BG2, fg=TXT3, anchor="w").pack(fill="x", padx=14, pady=4)

        # RIGHT – results grid
        right = tk.Frame(split, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        r_hdr = tk.Frame(right, bg=BG)
        r_hdr.pack(fill="x", padx=14, pady=(12,6))
        tk.Label(r_hdr, text="WYNIKI", font=(_SYS,8,"bold"),
                 bg=BG, fg=TXT3).pack(side="left")

        rsb = ttk.Scrollbar(right, orient="vertical")
        rsb.pack(side="right", fill="y")
        hsb2 = ttk.Scrollbar(right, orient="horizontal")
        hsb2.pack(side="bottom", fill="x")

        cols = ("Baza", "Status", "Błędy", "Ostrzeżenia", "KSeF", "Zakupy", "Sprzedaż", "Okres")
        self._tree = ttk.Treeview(right, columns=cols, show="headings",
                                   yscrollcommand=rsb.set,
                                   xscrollcommand=hsb2.set)
        rsb.config(command=self._tree.yview)
        hsb2.config(command=self._tree.xview)
        self._tree.pack(side="left", fill="both", expand=True, padx=(8,0))

        widths = {"Baza":260,"Status":80,"Błędy":60,"Ostrzeżenia":90,
                  "KSeF":60,"Zakupy":70,"Sprzedaż":70,"Okres":100}
        for c in cols:
            self._tree.heading(c, text=c, anchor="w")
            self._tree.column(c, width=widths.get(c,80), minwidth=40, anchor="w")

        self._tree.tag_configure("error",   foreground=ERR)
        self._tree.tag_configure("warning", foreground=WARN)
        self._tree.tag_configure("ok",      foreground=OK)
        self._tree.tag_configure("running", foreground=ACCENT)
        self._tree.tag_configure("pending", foreground=TXT2)

        self._tree.bind("<Double-Button-1>", self._on_tree_dbl)

        tk.Label(right, text="ℹ  Dwuklik na wierszu = otwórz w zakładce Pojedyncza baza",
                 font=FSM, bg=BG, fg=TXT3, anchor="w").pack(
                 side="bottom", fill="x", padx=8, pady=2)

    # ── file management ───────────────────────────────────────────────────
    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Wybierz bazy MDB",
            filetypes=[("Access Database","*.mdb *.MDB *.accdb"),
                       ("Wszystkie pliki","*.*")])
        for p in paths:
            if p not in [f[0] for f in self._files]:
                self._files.append((p, None, None))
                self._file_list.insert("end", os.path.basename(p))
        if self._files:
            self._run_btn.config(state="normal")

    def _remove_file(self, idx):
        if idx < 0 or idx >= len(self._files): return
        self._files.pop(idx)
        self._file_list.delete(idx)
        if not self._files: self._run_btn.config(state="disabled")

    def _clear_list(self):
        self._files.clear()
        self._file_list.delete(0, "end")
        self._tree.delete(*self._tree.get_children())
        self._results.clear()
        self._run_btn.config(state="disabled")
        self._export_btn.config(state="disabled")
        self._status.config(text="")

    # ── batch run ─────────────────────────────────────────────────────────
    def _run_all(self):
        if self._running: return
        self._running = True
        self._results.clear()
        self._tree.delete(*self._tree.get_children())
        self._export_btn.config(state="disabled")
        self._run_btn.config(state="disabled")

        month, year = None, None
        if self._period_var.get() == "month":
            try:
                month = MONTHS_PL.index(self._month_var.get()) + 1
                year  = int(self._year_var.get())
            except (ValueError, IndexError): pass

        # insert pending rows
        self._iids = []
        for path, _, _ in self._files:
            iid = self._tree.insert("", "end",
                values=(os.path.basename(path), "⏳ Czeka",
                        "–","–","–","–","–",
                        f"{MONTHS_PL[month-1]} {year}" if month else "Cały"),
                tags=("pending",))
            self._iids.append(iid)

        self._prog.config(maximum=len(self._files), value=0)
        self._prog.pack(side="left", padx=6)

        def worker():
            cfg = self._cfg_getter()
            for i, (path, _, _) in enumerate(self._files):
                iid = self._iids[i]
                self.after(0, lambda iid=iid: self._tree.item(iid,
                    values=(self._tree.item(iid,"values")[0],
                            "⟳ Sprawdzam","–","–","–","–","–",
                            self._tree.item(iid,"values")[7]),
                    tags=("running",)))
                self.after(0, lambda m=f"[{i+1}/{len(self._files)}] {os.path.basename(path)[:40]}…":
                           self._status.config(text=m, fg=TXT2))
                try:
                    res = analyze_mdb(path, month=month, year=year, cfg=cfg)
                    # tryb wsadowy nie otwiera popupów — nie trzymaj DataFrame'ów w RAM
                    res.pop("_data", None)
                    self._results.append((path, res))
                    s   = res.get("summary",{})
                    err = sum(1 for c in res["checks"] if c["kind"]=="error")
                    wrn = sum(1 for c in res["checks"] if c["kind"]=="warning")
                    if err:      tag, st = "error",   f"✗ {err} błędów"
                    elif wrn:    tag, st = "warning",  f"⚠ {wrn} ostrzeżeń"
                    else:        tag, st = "ok",       "✓ OK"
                    self.after(0, lambda iid=iid, tag=tag, st=st, s=s:
                               self._tree.item(iid,
                                   values=(self._tree.item(iid,"values")[0],
                                           st, err, wrn,
                                           s.get("ksef_total","?"),
                                           s.get("ksef_zakup","?"),
                                           s.get("ksef_sprz","?"),
                                           self._tree.item(iid,"values")[7]),
                                   tags=(tag,)))
                except Exception as e:
                    self._results.append((path, None))
                    msg = str(e)[:40]
                    self.after(0, lambda iid=iid, msg=msg:
                               self._tree.item(iid,
                                   values=(self._tree.item(iid,"values")[0],
                                           f"✗ Błąd: {msg}","–","–","–","–","–",
                                           self._tree.item(iid,"values")[7]),
                                   tags=("error",)))
                self.after(0, lambda i=i: self._prog.config(value=i+1))

            self.after(0, self._on_done)

        threading.Thread(target=worker, daemon=True).start()

    def _on_done(self):
        self._running = False
        self._run_btn.config(state="normal")
        self._prog.pack_forget()
        ok_n  = sum(1 for _,r in self._results if r and r.get("status")=="ok")
        err_n = sum(1 for _,r in self._results if r and r.get("status")=="error")
        wrn_n = sum(1 for _,r in self._results if r and r.get("status")=="warning")
        self._status.config(
            text=f"Gotowe: {ok_n}✓  {wrn_n}⚠  {err_n}✗  z {len(self._results)} baz",
            fg=ERR if err_n else (WARN if wrn_n else OK))
        if self._results:
            self._export_btn.config(state="normal")

    def _on_tree_dbl(self, e):
        """Dwuklik na wierszu → otwórz bazę w zakładce Pojedyncza baza."""
        item = self._tree.identify_row(e.y)
        if not item: return
        vals = self._tree.item(item, "values")
        # znajdź pełną ścieżkę na podstawie nazwy pliku (kolumna 0)
        bname = vals[0] if vals else ""
        path = next((p for p, *_ in self._files if os.path.basename(p) == bname), None)
        if not path and self._results:
            path = next((p for p, _ in self._results if os.path.basename(p) == bname), None)
        if not path or not self._on_open_single: return

        month, year = None, None
        if self._period_var.get() == "month":
            try:
                month = MONTHS_PL.index(self._month_var.get()) + 1
                year  = int(self._year_var.get())
            except (ValueError, IndexError): pass

        self._on_open_single(path, month, year)

    # ── batch export ──────────────────────────────────────────────────────
    def _export_batch(self):
        path = filedialog.asksaveasfilename(
            title="Zapisz raport zbiorczy",
            defaultextension=".xlsx",
            initialfile="KSeF_Raport_Zbiorczy.xlsx",
            filetypes=[("Excel","*.xlsx"),("Wszystkie pliki","*.*")])
        if not path: return
        try:
            _export_batch_excel(self._results, path)
            self._status.config(text=f"✓ Zapisano: {os.path.basename(path)}", fg=OK)
        except Exception as ex:
            import traceback
            self._status.config(text=f"✗ Błąd eksportu", fg=ERR)
            from tkinter import messagebox
            messagebox.showerror("Błąd eksportu Excel",
                                 f"{ex}\n\n{traceback.format_exc()[-600:]}")


def _export_batch_excel(results, save_path):
    """Zbiorczy raport Excel dla wielu baz."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import os, datetime

    wb = Workbook(); wb.remove(wb.active)

    def fill(h): return PatternFill("solid", fgColor=h)
    def font(h, bold=False, sz=9): return Font(name="Segoe UI", color=h, bold=bold, size=sz)
    thin = Side(style="thin", color="3d3d3d")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Arkusz 1: Podsumowanie zbiorcze ──────────────────────────────────
    ws = wb.create_sheet("Podsumowanie zbiorcze")
    ws.sheet_view.showGridLines = False

    ws.cell(1,1, f"KSeF Checker — Raport zbiorczy   |   {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws.cell(1,1).font = font("fb923c", bold=True, sz=12)
    ws.cell(1,1).fill = fill("1a1a1a")
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 24

    headers = ["Baza danych","Okres","Status","Błędy","Ostrzeżenia",
               "KSeF łącznie","Zakupy","Sprzedaż","Księga","VAT zakupy"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(2, j, h)
        c.font  = font("86efac", bold=True)
        c.fill  = fill("2a2a2a")
        c.border = brd

    col_w = [45,15,18,8,12,12,10,10,10,12]
    for j, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    kind_colors = {"ok":"86efac","warning":"fbbf24","error":"f87171"}

    for i, (path, res) in enumerate(results):
        row = i + 3
        bg = "222222" if i%2==0 else "1a1a1a"
        if res is None:
            for j in range(1, len(headers)+1):
                ws.cell(row,j,"" if j>1 else os.path.basename(path)).fill = fill(bg)
            ws.cell(row,3,"✗ Błąd odczytu").font = font("f87171")
            continue

        s    = res.get("summary",{})
        err  = sum(1 for c in res["checks"] if c["kind"]=="error")
        wrn  = sum(1 for c in res["checks"] if c["kind"]=="warning")
        stat = res.get("status","ok")
        fc   = kind_colors.get(stat,"f0f0f0")
        st   = "✓ OK" if stat=="ok" else (f"✗ {err} błędów" if stat=="error" else f"⚠ {wrn} ostrzeżeń")

        vals = [os.path.basename(path), res.get("period","?"), st,
                err or "", wrn or "",
                s.get("ksef_total","?"), s.get("ksef_zakup","?"),
                s.get("ksef_sprz","?"),  s.get("ksiega","?"), s.get("vatzakupy","?")]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v)
            c.fill   = fill(bg)
            c.border = brd
            c.font   = font(fc if j==3 else "f0f0f0", bold=(j==3))

    # ── Arkusz 2: Wszystkie błędy (zbiorczo) ─────────────────────────────
    ws2 = wb.create_sheet("Wszystkie błędy")
    ws2.sheet_view.showGridLines = False
    hdrs2 = ["Baza","Rodzaj","Sprawdzenie","Liczba","Szczegóły"]
    for j, h in enumerate(hdrs2, 1):
        c = ws2.cell(1, j, h)
        c.font = font("86efac", bold=True)
        c.fill = fill("2a2a2a")
        c.border = brd
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 52
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 60

    row = 2
    for path, res in results:
        if not res: continue
        for chk in res["checks"]:
            if chk["kind"] not in ("error","warning"): continue
            bg = "222222" if row%2==0 else "1a1a1a"
            fc = "f87171" if chk["kind"]=="error" else "fbbf24"
            vals = [os.path.basename(path), chk["kind"].upper(),
                    chk["title"], len(chk.get("rows",[])),
                    chk.get("detail","")]
            for j, v in enumerate(vals, 1):
                c = ws2.cell(row, j, v)
                c.fill = fill(bg)
                c.border = brd
                c.font = font(fc if j in (2,3) else "a0a0a0")
            row += 1

    # ── Arkusz 3: Do sprawdzenia (checklisty per firma) ───────────────────
    ws3 = wb.create_sheet("✉ Do sprawdzenia")
    ws3.sheet_view.showGridLines = False
    ws3.cell(1,1,"Firma").font  = font("86efac", bold=True); ws3.cell(1,1).fill = fill("2a2a2a")
    ws3.cell(1,2,"Problem").font = font("86efac", bold=True); ws3.cell(1,2).fill = fill("2a2a2a")
    ws3.cell(1,3,"Pilność").font = font("86efac", bold=True); ws3.cell(1,3).fill = fill("2a2a2a")
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 12

    row = 2
    for path, res in results:
        if not res: continue
        name = os.path.splitext(os.path.basename(path))[0][:35]
        for chk in res["checks"]:
            if chk["kind"] == "ok": continue
            bg = "222222" if row%2==0 else "1a1a1a"
            fc = "f87171" if chk["kind"]=="error" else "fbbf24"
            ws3.cell(row,1,name).font  = font("f0f0f0"); ws3.cell(row,1).fill = fill(bg)
            ws3.cell(row,2,chk["title"]).font = font(fc); ws3.cell(row,2).fill = fill(bg)
            ws3.cell(row,3,chk["kind"].upper()).font = font(fc,bold=True); ws3.cell(row,3).fill = fill(bg)
            row += 1

    wb.save(save_path)


# ── okno pobierania aktualizacji ─────────────────────────────────────────────
class UpdateDialog(tk.Toplevel):
    """
    Fazy:
      1. Pobieranie  — pasek postępu + MB/MB
      2. Instalacja  — indeterminate + "Instaluję..."
      3. Gotowe      — komunikat + przycisk "Uruchom ponownie"
    Aplikacja NIE zamyka się dopóki użytkownik nie kliknie "Uruchom ponownie".
    """
    def __init__(self, parent, app_root, version, installer_url):
        super().__init__(parent)
        self.title(f"Aktualizacja do {version}")
        self.geometry("460x200")
        self.configure(bg=BG)
        self.resizable(False, False)
        self.grab_set()
        self._app_root      = app_root
        self._version       = version
        self._installer_url = installer_url
        self._cancelled     = False
        self._build()
        self.lift(); self.focus_force()
        threading.Thread(target=self._download, daemon=True).start()

    def _build(self):
        tk.Label(self, text=f"Aktualizacja  →  KSeF Checker {self._version}",
                 font=FMED, bg=BG, fg=TXT, pady=16).pack()

        style = ttk.Style(self)
        style.configure("U.Horizontal.TProgressbar",
                        troughcolor=BG3, background=ACCENT,
                        bordercolor=BG3, lightcolor=ACCENT, darkcolor=ACCENT)
        self._prog = ttk.Progressbar(self, style="U.Horizontal.TProgressbar",
                                      length=400, mode="determinate")
        self._prog.pack(padx=28)

        self._lbl = tk.Label(self, text="Łączę z serwerem…",
                             font=FSM, bg=BG, fg=TXT2)
        self._lbl.pack(pady=10)

        self._btn = tk.Button(self, text="Anuluj", font=FSM,
                              bg=BG3, fg=TXT, activebackground=BG4,
                              relief="flat", padx=14, pady=5,
                              command=self._cancel)
        self._btn.pack()

    # ── faza 1: pobieranie ────────────────────────────────────────────────
    def _download(self):
        import urllib.request, tempfile
        try:
            tmp = tempfile.NamedTemporaryFile(suffix=".exe", delete=False)
            tmp_path = tmp.name
            tmp.close()

            def on_progress(count, block, total):
                if self._cancelled: raise Exception("Anulowano")
                if total > 0:
                    pct  = min(100, count * block * 100 // total)
                    done = count * block / 1048576
                    tot  = total / 1048576
                    self.after(0, lambda p=pct, d=done, t=tot: (
                        self._prog.config(value=p),
                        self._lbl.config(text=f"Pobieranie…  {d:.1f} MB / {t:.1f} MB")))

            urllib.request.urlretrieve(self._installer_url, tmp_path, on_progress)
            self.after(0, lambda: self._start_install(tmp_path))

        except Exception as e:
            if not self._cancelled:
                self.after(0, lambda: self._lbl.config(text=f"Błąd: {e}", fg=ERR))

    # ── faza 2: instalacja w tle ──────────────────────────────────────────
    def _start_install(self, installer_path):
        self._lbl.config(
            text="Instaluję…  (może pojawić się okno uprawnień administracyjnych)", fg=TXT2)
        self._prog.config(mode="indeterminate", value=0)
        self._prog.start(12)
        self._btn.config(state="disabled", text="Proszę czekać…")

        def run():
            try:
                import ctypes
                import ctypes.wintypes as wt

                # ShellExecuteEx z runas = właściwe żądanie UAC + czekanie na
                # FAKTYCZNY proces instalatora (nie pośredni launcher)
                class SHELLEXECUTEINFOW(ctypes.Structure):
                    _fields_ = [
                        ("cbSize",         wt.DWORD),
                        ("fMask",          wt.ULONG),
                        ("hwnd",           wt.HWND),
                        ("lpVerb",         wt.LPCWSTR),
                        ("lpFile",         wt.LPCWSTR),
                        ("lpParameters",   wt.LPCWSTR),
                        ("lpDirectory",    wt.LPCWSTR),
                        ("nShow",          ctypes.c_int),
                        ("hInstApp",       wt.HINSTANCE),
                        ("lpIDList",       ctypes.c_void_p),
                        ("lpClass",        wt.LPCWSTR),
                        ("hkeyClass",      wt.HKEY),
                        ("dwHotKey",       wt.DWORD),
                        ("hIconOrMonitor", wt.HANDLE),
                        ("hProcess",       wt.HANDLE),
                    ]

                SEE_MASK_NOCLOSEPROCESS = 0x00000040
                sei = SHELLEXECUTEINFOW()
                sei.cbSize       = ctypes.sizeof(sei)
                sei.fMask        = SEE_MASK_NOCLOSEPROCESS
                sei.hwnd         = None
                sei.lpVerb       = "runas"          # wymuś UAC / admin
                sei.lpFile       = installer_path
                # instalator i tak ubija stary proces (taskkill) przed podmianą
                sei.lpParameters = "/VERYSILENT /NORESTART"
                sei.lpDirectory  = None
                sei.nShow        = 0                # SW_HIDE
                sei.hProcess     = None

                ok = ctypes.windll.shell32.ShellExecuteExW(ctypes.byref(sei))
                if not ok or not sei.hProcess:
                    err = ctypes.windll.kernel32.GetLastError()
                    raise RuntimeError(
                        f"Nie udało się uruchomić instalatora "
                        f"(ShellExecuteEx error {err}). "
                        f"Pobierz instalator ręcznie ze strony GitHub.")

                ctypes.windll.kernel32.CloseHandle(sei.hProcess)
                # Zamknij aplikację OD RAZU — działający exe blokuje własny plik
                # i instalator nie mógłby go podmienić. Po instalacji użytkownik
                # otwiera program ponownie (ręcznie = uruchomi się bez uprawnień admina).
                self.after(0, self._exit_for_update)
            except Exception as e:
                self.after(0, lambda: self._lbl.config(
                    text=f"Błąd instalacji: {e}", fg=ERR))

        threading.Thread(target=run, daemon=True).start()

    # ── faza 3: zamknięcie na czas instalacji ─────────────────────────────
    def _exit_for_update(self):
        self._prog.stop()
        self._prog.config(mode="determinate", value=100)
        self._lbl.config(
            text=f"✓  Instaluję {self._version} — za chwilę otwórz program ponownie",
            fg=OK)
        # krótka chwila na pokazanie komunikatu, potem twarde wyjście
        # (zwalnia KSeF_Checker.exe zanim instalator dojdzie do kopiowania)
        self.after(1500, lambda: os._exit(0))

    def _cancel(self):
        self._cancelled = True
        self.destroy()


# ── historia analiz ───────────────────────────────────────────────────────────
# ── historia analiz ───────────────────────────────────────────────────────────
class HistoryWindow(tk.Toplevel):
    def __init__(self, parent, on_load):
        super().__init__(parent)
        self.title("Historia analiz — KSeF Checker")
        self.geometry("740x560")
        self.configure(bg=BG)
        self.resizable(True, True)
        self._parent   = parent
        self._on_load  = on_load
        self._history  = cfg_module.load_history()
        self._build()
        self.lift(); self.focus_force()

    def _build(self):
        for w in self.winfo_children(): w.destroy()

        # nagłówek
        hdr = tk.Frame(self, bg=BG2, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🕐  Historia analiz", font=FMED,
                 bg=BG2, fg=ACCENT, padx=16).pack(side="left")
        tk.Label(hdr, text=f"{len(self._history)} zapisanych wyników",
                 font=FSM, bg=BG2, fg=TXT3).pack(side="left")
        if self._history:
            tk.Button(hdr, text="Wyczyść wszystko", font=FSM,
                      bg=BG3, fg=ERR, activebackground=BG4,
                      relief="flat", padx=10, pady=4,
                      command=self._clear_all).pack(side="right", padx=12)

        if not self._history:
            tk.Label(self,
                     text="Brak historii.\n\nUruchom analizę — wyniki zostaną tu zapisane automatycznie.",
                     font=FB, bg=BG, fg=TXT3, justify="center").pack(expand=True)
            return

        # przewijalna lista wpisów
        cv = tk.Canvas(self, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=cv.yview)
        cv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        cv.pack(fill="both", expand=True, padx=8, pady=8)
        inner = tk.Frame(cv, bg=BG)
        wid = cv.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.bind("<Configure>",    lambda e: cv.itemconfig(wid, width=e.width))
        cv.bind("<MouseWheel>",   lambda e: cv.yview_scroll(-1*(e.delta//120), "units"))

        kc = {"ok": (OK,"✓"), "warning": (WARN,"⚠"), "error": (ERR,"✗")}

        for i, entry in enumerate(self._history):
            col, icon = kc.get(entry["status"], (TXT2, "·"))
            n_err = sum(1 for c in entry.get("checks",[]) if c["kind"]=="error")
            n_wrn = sum(1 for c in entry.get("checks",[]) if c["kind"]=="warning")
            s     = entry.get("summary", {})

            card = tk.Frame(inner, bg=BG2, cursor="hand2",
                            highlightbackground=col if entry["status"]!="ok" else BORDER,
                            highlightthickness=1)
            card.pack(fill="x", pady=2)

            # lewy pasek koloru
            tk.Frame(card, bg=col if entry["status"]!="ok" else BG4,
                     width=4).pack(side="left", fill="y")

            # treść
            content = tk.Frame(card, bg=BG2)
            content.pack(side="left", fill="both", expand=True, padx=10, pady=8)

            row1 = tk.Frame(content, bg=BG2)
            row1.pack(fill="x")
            tk.Label(row1, text=icon, font=(_SYS,10,"bold"),
                     bg=BG2, fg=col, width=2).pack(side="left")
            tk.Label(row1, text=entry["filename"], font=(_SYS,10,"bold"),
                     bg=BG2, fg=TXT, anchor="w").pack(side="left", fill="x", expand=True)
            ts = entry.get("timestamp","")[:16].replace("T", "  ")
            tk.Label(row1, text=ts, font=("Consolas",8),
                     bg=BG2, fg=TXT3).pack(side="right")

            row2 = tk.Frame(content, bg=BG2)
            row2.pack(fill="x", pady=(3,0))
            tk.Label(row2, text=f"📅  {entry.get('period','?')}",
                     font=FSM, bg=BG2, fg=TXT2).pack(side="left", padx=(20,10))
            if n_err:
                tk.Label(row2, text=f" ✗ {n_err} ", font=("Consolas",8,"bold"),
                         bg=ERR, fg="#18181b").pack(side="left", padx=2)
            if n_wrn:
                tk.Label(row2, text=f" ⚠ {n_wrn} ", font=("Consolas",8,"bold"),
                         bg=WARN, fg="#18181b").pack(side="left", padx=2)
            if not n_err and not n_wrn:
                tk.Label(row2, text="Wszystko OK", font=FSM,
                         bg=BG2, fg=OK).pack(side="left", padx=2)
            # statystyki KSeF
            ksef_total = s.get("ksef_total","?")
            tk.Label(row2, text=f"KSeF: {ksef_total}",
                     font=FSM, bg=BG2, fg=TXT3).pack(side="right", padx=8)

            # przyciski
            btn_col = tk.Frame(card, bg=BG2)
            btn_col.pack(side="right", padx=8, pady=8)
            tk.Button(btn_col, text="  Otwórz  ", font=FSM,
                      bg=ACCENT, fg="#18181b",
                      activebackground="#ea6d05", activeforeground="#18181b",
                      relief="flat", padx=8, pady=4,
                      command=lambda e=entry: self._load(e)).pack(fill="x", pady=(0,4))
            tk.Button(btn_col, text="  Usuń  ", font=FSM,
                      bg=BG3, fg=TXT2,
                      activebackground=BG4, activeforeground=ERR,
                      relief="flat", padx=8, pady=4,
                      command=lambda idx=i: self._delete(idx)).pack(fill="x")

            # dwuklik na karcie też otwiera
            for w in card.winfo_children() + content.winfo_children() + \
                     row1.winfo_children() + row2.winfo_children():
                w.bind("<Double-Button-1>", lambda e, en=entry: self._load(en))
            card.bind("<Double-Button-1>", lambda e, en=entry: self._load(en))

        tk.Label(self, text="ℹ  Dwuklik lub przycisk Otwórz = wczytaj wyniki bez ponownej analizy",
                 font=FSM, bg=BG, fg=TXT3, anchor="w").pack(
                 fill="x", padx=12, pady=4)

    def _load(self, entry):
        self.destroy()
        self._on_load(entry)

    def _delete(self, idx):
        cfg_module.delete_history_entry(idx)
        self._history = cfg_module.load_history()
        self._build()

    def _clear_all(self):
        import os
        try: os.remove(cfg_module.HISTORY_FILE)
        except Exception: pass
        self._history = []
        self._build()


# ── help window ───────────────────────────────────────────────────────────────
class HelpWindow(tk.Toplevel):
    """Okno pomocy z instrukcją użytkownika."""

    # ── treść instrukcji ──────────────────────────────────────────────────────
    _SECTIONS = [
        ("Opis programu", """
KSeF Checker to narzędzie do weryfikacji spójności danych między bazą KSeF a bazą księgową programu Mała Księgowość (plik .mdb). Program automatycznie wykrywa rozbieżności między fakturami pobranymi z Krajowego Systemu e-Faktur a zapisami w:

  • KSIEGA — księdze przychodów i rozchodów
  • VATZAKUPY — rejestrze VAT zakupów
  • VATSPRZEDAZ — rejestrze VAT sprzedaży

Program nie modyfikuje żadnych danych — działa wyłącznie w trybie odczytu.
"""),
        ("Instalacja", """
WYMAGANIA SYSTEMOWE
  • Windows 10 / 11 (64-bit)
  • Microsoft Access Database Engine 2016 (64-bit)

WARIANT A — z Access Database Engine (dla nowych instalacji)
Pobierz i uruchom KSeF_Checker_Installer_z_AccessEngine.exe. Instalator automatycznie zainstaluje sterownik Microsoft Access jeśli nie jest jeszcze obecny.

WARIANT B — bez Access Database Engine (jeśli sterownik już jest)
Pobierz i uruchom KSeF_Checker_Installer.exe.

Uwaga: Jeśli masz zainstalowany pakiet Microsoft Office (32-bit), sterownik Access musi być w tej samej wersji bitowej co Office. W razie konfliktu skontaktuj się z administratorem IT.
"""),
        ("Przeprowadzenie analizy", """
KROK 1 — Otwórz plik bazy danych
Kliknij przycisk 📂 Otwórz plik MDB i wskaż plik bazy danych programu Mała Księgowość (rozszerzenie .mdb).

KROK 2 — Wybierz okres analizy

  Cały      — sprawdzenie całej bazy (wszystkich dostępnych danych)
  Miesiąc   — analiza konkretnego miesiąca (najczęstszy tryb)
  Kwartał   — analiza kwartalnych deklaracji VAT (Q1–Q4)

KROK 3 — Uruchom analizę
Kliknij ▶ Sprawdź. Program połączy się z bazą, wczyta tabele i przeprowadzi 19 sprawdzeń. Czas analizy: zazwyczaj 2–10 sekund.

KROK 4 — Przeglądaj wyniki
  ✓ Karty zielone   — brak nieprawidłowości
  ⚠ Karty żółte    — ostrzeżenia wymagające weryfikacji
  ✗ Karty czerwone  — błędy wymagające korekty

Kliknij kartę, aby zobaczyć szczegóły w panelu po prawej.
"""),
        ("Filtry i sortowanie", """
FILTRY KATEGORII
Na górze lewego panelu znajdują się przyciski filtrów:
  [ Wszystkie ] [ Kompletność ] [ Kwoty ] [ Daty ]
  [ Jakość ] [ Korekty ] [ Compliance ]

Kliknięcie kategorii pokazuje tylko jej sprawdzenia i automatycznie zaznacza pierwszy błąd/ostrzeżenie z tej grupy.

SORTOWANIE TABELI
W panelu szczegółów (prawa strona):
  • Kliknij nagłówek kolumny — sortuje rosnąco (↑), ponowne kliknięcie malejąco (↓)
  • Przyciski „Sortuj:" obok licznika wierszy — szybki dostęp do sortowania po dacie lub kwocie
"""),
        ("Popup szczegółów faktury", """
DWUKLIK NA WIERSZU otwiera okno ze szczegółowym porównaniem danych z trzech źródeł:

  ┌──────────┐          ┌───────────────┐
  │  KSeF    │   ⟷    │  Rejestr VAT  │
  │          │  Δ Brutto │              │
  │  Netto   │  Δ Netto  │  Netto        │
  │ 1 230,00 │  Δ VAT    │    189.35     │
  │  Brutto  │          │  Brutto       │
  │ 1 512,90 │          │  1 462.90     │
  └──────────┘          └───────────────┘

Wartości Δ zaznaczone na czerwono oznaczają rozbieżność.
Wartość 0,00 oznacza zgodność.

Dla sprawdzeń z KSIEGA wyświetlana jest również karta Księga z kwotą kosztów i procentem odliczenia.
"""),
        ("Oznaczanie faktur", """
Podczas przeglądania wyników możesz oznaczać faktury wymagające poprawki w programie księgowym.

JAK OZNACZAĆ
  • Prawy klik na wierszu → wiersz zmienia kolor na zielony (oznaczono)
  • Prawy klik ponownie → odznaczenie
  • Przycisk „Odznacz wszystkie" — kasuje wszystkie oznaczenia danego sprawdzenia

LICZNIK
Nad tabelą widoczny jest licznik ✓ N zaznaczonych aktualizowany na bieżąco.

TRWAŁOŚĆ
Oznaczenia są zapisywane w pliku marks.json obok programu. Przypisane są do konkretnego pliku MDB i konkretnego sprawdzenia — przeżywają zamknięcie programu.

TYPOWY PRZEPŁYW PRACY
  1. Uruchom analizę dla miesiąca
  2. Otwórz program Mała Księgowość obok
  3. W KSeF Checker oznaczaj faktury prawym klikiem w miarę ich poprawiania
  4. Po zakończeniu uruchom analizę ponownie — liczba błędów powinna spaść do zera
"""),
        ("Analiza wielu baz (Batch)", """
Zakładka Wiele baz (batch) pozwala sprawdzić kilka firm jednocześnie.

JAK UŻYWAĆ
  1. Kliknij ＋ Dodaj bazy MDB — wybierz wiele plików naraz (Ctrl+klik)
  2. Ustaw wspólny okres analizy
  3. Kliknij ▶ Sprawdź wszystkie
  4. Wyniki pojawią się w tabeli z podsumowaniem per firma

EKSPORT ZBIORCZY
Kliknij ↓ Excel zbiorczy — plik zawiera:
  • Arkusz Podsumowanie zbiorcze — jedna linia per firma
  • Arkusz Wszystkie błędy — pełna lista problemów ze wszystkich baz
  • Arkusz ✉ Do sprawdzenia — checklist do przekazania klientom

OTWARCIE FIRMY W TRYBIE SZCZEGÓŁOWYM
Dwuklik na wierszu w tabeli zbiorczej otwiera tę bazę w zakładce Pojedyncza baza i automatycznie uruchamia analizę.
"""),
        ("Historia analiz", """
Kliknij 🕐 Historia w nagłówku programu, aby zobaczyć poprzednie analizy.

  • Wyniki są zapisywane automatycznie po każdej analizie
  • Przechowywanych jest do 40 ostatnich wpisów
  • Dla tej samej bazy i okresu poprzedni wynik jest nadpisywany

WCZYTYWANIE Z HISTORII
Kliknij Otwórz przy wybranym wpisie — program wczyta zapisane wyniki bez ponownej analizy bazy. Przydatne gdy baza MDB jest niedostępna (np. klient zdalny).
"""),
        ("Eksport do Excel", """
Kliknij ↓ Excel w pasku narzędzi po zakończeniu analizy.

ZAWARTOŚĆ PLIKU

Arkusz „Raport" — tabela problemów:
  • Typ (BŁĄD / OSTRZEŻENIE)
  • Sprawdzenie — nazwa wykrytego problemu
  • Liczba rekordów — ile faktur dotyczy
  • Szczegóły — sumy rozbieżności
  • Wyjaśnienie — podstawa prawna

Arkusz „Szczegóły" — wszystkie wiersze z nieprawidłowościami z pełnymi danymi faktur (numery, daty, kwoty, kontrahenci).

Nazwa pliku jest generowana automatycznie: KSeF_NazwaFirmy_Maj_2024.xlsx
"""),
        ("Ustawienia", """
Kliknij ⚙ Ustawienia w prawym górnym rogu.

ZAKŁADKA „SPRAWDZENIA"
Włącz lub wyłącz poszczególne sprawdzenia. Przyciski:
  • Zaznacz wszystkie — włącz wszystkie
  • Odznacz wszystkie — wyłącz wszystkie
  • Przywróć domyślne — wróć do ustawień fabrycznych

ZAKŁADKA „PARAMETRY"

  Tolerancja różnic kwotowych (domyślnie: 0,05 zł)
    Różnice poniżej tej wartości są ignorowane (zaokrąglenia groszy).

  Próg ostrzeżenia „brak KSeF" (domyślnie: 15%)
    Przy jakim odsetku faktur bez numeru KSeF wyświetlić ostrzeżenie.

  Dni bufora dla „cały zakres" (domyślnie: 3 dni)
    Ile ostatnich dni ignorować jako „za świeże do zaksięgowania"
    (faktury pobrane z KSeF mogą jeszcze nie być zaksięgowane).
"""),
        ("Aktualizacje", """
Program automatycznie sprawdza dostępność aktualizacji przy każdym uruchomieniu (raz dziennie).

Gdy dostępna jest nowa wersja, w górnej części okna pojawia się baner:
  🔔 Dostępna nowa wersja v1.x.x  —  [  Aktualizuj  ]

PROCES AKTUALIZACJI
  1. Kliknij Aktualizuj
  2. Postęp pobierania jest widoczny na pasku
  3. Może pojawić się okno UAC „Czy pozwolić tej aplikacji..." — kliknij Tak
  4. Po instalacji kliknij 🔄 Uruchom ponownie
  5. Program automatycznie uruchamia nową wersję

Program nie zamyka się podczas pobierania — można pracować do momentu kliknięcia „Uruchom ponownie".
"""),
        ("Opis sprawdzeń", """
KOMPLETNOŚĆ

  ✗ Niezaksięgowane zakupy z KSeF
    Faktury zakupowe z KSeF bez żadnego wpisu w KSIEGA ani VATZAKUPY.
    Co zrobić: zaksięgować fakturę w Mała Księgowość.

  ✗ Niezaksięgowana sprzedaż z KSeF
    Faktury sprzedażowe z KSeF bez wpisu w KSIEGA lub VATSPRZEDAZ.

  ✗ Niespójność KSIEGA ↔ VATZAKUPY
    Faktura jest w księdze ale nie ma jej w rejestrze VAT (lub odwrotnie).
    Faktura powinna być jednocześnie w obu miejscach.

─────────────────────────────────────────────────────────────

KWOTY

  ✗/⚠ Niezgodności kwotowe — KSeF vs VATZAKUPY
    Porównuje netto, VAT i brutto między KSeF a rejestrem VAT zakupów.
    Kolumny Δ Netto / Δ VAT / Δ Brutto pokazują różnicę (0,00 = zgodne).
    Obsługuje faktury ZAKUP50 (50% odliczenie) i faktury zwolnione z VAT.

  ⚠ Niezgodności kwotowe — KSeF vs KSIEGA
    Porównuje netto z KSeF z sumą kosztów w księdze.
    Uwzględnia kolumnę % kosztu (np. 75% dla samochodu).

  ✗/⚠ Niezgodności kwotowe sprzedaży — KSeF vs VATSPRZEDAZ
    Porównuje kwoty faktur sprzedażowych.

─────────────────────────────────────────────────────────────

DATY

  ⚠ Zakupy ujęte poza dopuszczalnym oknem
    Art. 86 ust. 11 ustawy o VAT pozwala ująć fakturę w miesiącu
    wystawienia lub 3 kolejnych miesiącach (M+1, M+2, M+3).
    Flagowane są tylko ujęcia WSTECZ lub po upływie M+3.
    Ujęcie w M+1/M+2/M+3 jest legalne i NIE jest flagowane.

  ⚠ Sprzedaż w złym miesiącu (VATSPRZEDAZ)
    Faktura sprzedaży ujęta w innym miesiącu niż data sprzedaży.

─────────────────────────────────────────────────────────────

JAKOŚĆ

  ⚠ Zakupy bez numeru KSeF
    Faktury w rejestrze VAT bez numeru KSeF (ręczne, zagraniczne, noty).
    Ostrzeżenie gdy odsetek przekroczy próg z Ustawień (domyślnie 15%).

  ⚠ Wpisy tylko w KSIEGA (ZUS, opłaty, wynagrodzenia)
    Wydatki w księdze bez odpowiednika w rejestrze VAT — ZUS, opłaty
    bankowe, wynagrodzenia, amortyzacja. Normalny rodzaj wydatku,
    wyświetlany do weryfikacji kompletności.

  ✗ Duplikaty w VATZAKUPY
    Ten sam numer KSeF zaksięgowany więcej niż raz — zawyża VAT
    i powoduje błąd w pliku JPK_VAT.

  ✗ Duplikaty w VATSPRZEDAZ
    Jak wyżej, dla rejestru sprzedaży.

─────────────────────────────────────────────────────────────

KOREKTY

  ✗ Niezaksięgowane korekty zakupów
    Faktury korygujące (ujemne) z KSeF bez zaksięgowania.
    Niezaksięgowana korekta zawyża odliczony VAT i koszty.

  ✗ Niezaksięgowane korekty sprzedaży
    Analogicznie dla faktur korygujących sprzedaż.

─────────────────────────────────────────────────────────────

COMPLIANCE

  ✗ Błędny NIP kontrahenta
    NIP sprzedawcy w KSeF różni się od NIP w rejestrze VAT.
    Niezgodny NIP powoduje odrzucenie pliku JPK przez urząd.

  ✗ Przekroczony termin odliczenia VAT (>3 miesiące)
    Czas między datą wystawienia a datą ujęcia przekracza 3 miesiące
    kalendarzowe (art. 86 ust. 11). Kolumna „Miesięcy po terminie"
    pokazuje skalę przekroczenia.

  ⚠ Split payment (MPP) — faktury ≥ 15 000 zł
    Faktury bez oznaczenia MPP powyżej progu.
    UWAGA: MPP jest wymagane wyłącznie dla towarów z Załącznika 15
    do ustawy o VAT — weryfikuj ręcznie czy obowiązek dotyczy faktury.

  ⚠ Faktury walutowe (nie-PLN)
    Faktury w walutach obcych wymagające przeliczenia na PLN wg kursu
    NBP z dnia poprzedzającego wystawienie (art. 31a ustawy o VAT).

  ⚠ Ten sam NIP, różne nazwy firmy
    Możliwe literówki, zmiany nazwy firmy lub błędy przy wprowadzaniu.
"""),
        ("Najczęstsze pytania", """
P: Program wyświetla błąd „Brak sterownika Microsoft Access"
O: Zainstaluj sterownik ze strony Microsoft lub użyj instalatora
   z sufiksem _z_AccessEngine. Jeśli masz 32-bitowy Microsoft Office,
   musisz użyć 32-bitowej wersji sterownika.

─────────────────────────────────────────────────────────────

P: Faktura jest w KSeF ale program pokazuje ją jako niezaksięgowaną
O: Sprawdź czy numer KSeF w Mała Księgowość zgadza się dokładnie z numerem
   w KSeF (bez dodatkowych spacji). Możliwa przyczyna to literówka lub
   brak numeru KSeF przy ręcznym wprowadzaniu faktury.

─────────────────────────────────────────────────────────────

P: Program pokazuje stare korekty sprzed roku jako niezaksięgowane
O: Sprawdzenie korekt działa w zakresie wybranego okresu. Jeśli stare
   korekty nie powinny być księgowane, możesz wyłączyć to sprawdzenie
   w Ustawieniach.

─────────────────────────────────────────────────────────────

P: Część faktur ma status „zły miesiąc" choć je zaksięgowałem w M+1
O: Ujęcie faktury zakupowej w M+1, M+2 lub M+3 jest LEGALNE i program
   tego nie flaguje. Jeśli widzisz ostrzeżenie, faktura jest ujęta albo
   wstecz (przed miesiącem wystawienia) albo po upływie 3 miesięcy.

─────────────────────────────────────────────────────────────

P: Co znaczy „Δ Brutto = 50,00 zł"?
O: Kwota brutto w KSeF różni się od kwoty w rejestrze VAT o 50 zł.
   Może to być: błędna wartość przy ręcznym wpisie, zaokrąglenie lub
   zaksięgowanie innej faktury pod tym numerem KSeF.

─────────────────────────────────────────────────────────────

P: Gdzie są zapisywane ustawienia i historia?
O: W folderze instalacji (domyślnie C:\\Program Files\\KSeF Checker\\):
   • settings.json — ustawienia programu
   • history.json  — historia analiz (maks. 40 wpisów)
   • marks.json    — oznaczenia faktur „do sprawdzenia"
"""),
    ]

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Pomoc — KSeF Checker")
        self.geometry("1020x680")
        self.minsize(700, 480)
        self.configure(bg=BG)
        self.resizable(True, True)
        self._build()
        self.lift(); self.focus_force()

    def _build(self):
        # nagłówek
        hdr = tk.Frame(self, bg=BG2)
        hdr.pack(fill="x")
        tk.Label(hdr, text="❓  Pomoc — KSeF Checker", font=FMED,
                 bg=BG2, fg=ACCENT, padx=16, pady=10).pack(side="left")
        tk.Label(hdr, text=f"v{VERSION}", font=FSM,
                 bg=BG2, fg=TXT3, padx=4).pack(side="left")
        tk.Button(hdr, text="✕", font=FSM, bg=BG2, fg=TXT3,
                  activebackground=BG3, relief="flat", padx=10, pady=8,
                  cursor="hand2", command=self.destroy).pack(side="right", padx=8)
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        # split: lewy TOC + prawy content
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True)

        # ── lewy panel: spis treści ───────────────────────────────────────
        left = tk.Frame(body, bg=BG2, width=210)
        left.pack(side="left", fill="y"); left.pack_propagate(False)
        tk.Frame(body, bg=BORDER, width=1).pack(side="left", fill="y")

        tk.Label(left, text="SPIS TREŚCI", font=(_SYS,8,"bold"),
                 bg=BG2, fg=TXT3, padx=14, pady=10).pack(anchor="w")

        self._toc_btns = []
        for i, (title, _) in enumerate(self._SECTIONS):
            b = tk.Button(left, text=title, font=FSM,
                          bg=BG2, fg=TXT2, anchor="w",
                          activebackground=BG3, activeforeground=TXT,
                          relief="flat", padx=14, pady=5, cursor="hand2",
                          command=lambda idx=i: self._goto(idx))
            b.pack(fill="x")
            b.bind("<Enter>", lambda e, _b=b: _b.config(fg=TXT))
            b.bind("<Leave>", lambda e, _b=b, _i=i: _b.config(
                fg=ACCENT if self._active == _i else TXT2))
            self._toc_btns.append(b)

        # ── prawy panel: treść ────────────────────────────────────────────
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(right, orient="vertical")
        vsb.pack(side="right", fill="y")

        self._txt = tk.Text(right, bg=BG, fg=TXT, font=FS,
                            relief="flat", bd=0, padx=20, pady=14,
                            wrap="word", cursor="arrow",
                            yscrollcommand=vsb.set,
                            state="disabled", spacing1=2, spacing3=4)
        self._txt.pack(fill="both", expand=True)
        vsb.config(command=self._txt.yview)

        self._txt.bind("<MouseWheel>",
                       lambda e: self._txt.yview_scroll(-1*(e.delta//120),"units"))

        # tagi formatowania
        self._txt.tag_configure("h1",    font=(_SYS,15,"bold"), foreground=ACCENT,
                                spacing1=18, spacing3=6)
        self._txt.tag_configure("h2",    font=(_SYS,11,"bold"), foreground=TXT,
                                spacing1=14, spacing3=4)
        self._txt.tag_configure("sep",   foreground=TXT3)
        self._txt.tag_configure("mono",  font=(FMONO,9),        foreground=TXT2)
        self._txt.tag_configure("note",  foreground=WARN)
        self._txt.tag_configure("ok",    foreground=OK)
        self._txt.tag_configure("err",   foreground=ERR)
        self._txt.tag_configure("body",  font=FS, foreground=TXT,  spacing1=1)

        # wypełnij treść
        self._active = 0
        self._section_marks = []
        self._render_all()
        self._highlight_toc(0)

        # śledź widoczną sekcję przy scrollowaniu
        vsb.config(command=self._on_scroll)

    def _on_scroll(self, *args):
        self._txt.yview(*args)
        self._update_active_section()

    def _update_active_section(self):
        # sprawdź która sekcja jest aktualnie widoczna
        first_visible = self._txt.index("@0,0")
        for i, mark in reversed(list(enumerate(self._section_marks))):
            if self._txt.compare(mark, "<=", first_visible):
                if i != self._active:
                    self._active = i
                    self._highlight_toc(i)
                return

    def _highlight_toc(self, idx):
        for i, b in enumerate(self._toc_btns):
            b.config(bg=BG3 if i == idx else BG2,
                     fg=ACCENT if i == idx else TXT2)

    def _goto(self, idx):
        self._active = idx
        self._highlight_toc(idx)
        mark = self._section_marks[idx]
        self._txt.see(mark)

    def _render_all(self):
        self._txt.config(state="normal")
        self._txt.delete("1.0", "end")
        self._section_marks.clear()

        for i, (title, content) in enumerate(self._SECTIONS):
            # zapisz markę sekcji
            mark = f"sec_{i}"
            self._txt.mark_set(mark, "end-1c")
            self._txt.mark_gravity(mark, "left")
            self._section_marks.append(mark)

            # nagłówek sekcji
            if i > 0:
                self._txt.insert("end", "\n")
            self._txt.insert("end", f"{title}\n", "h1")

            # treść linii po linii
            for line in content.strip().split("\n"):
                stripped = line.strip()
                if stripped.startswith("─────"):
                    self._txt.insert("end", f"{stripped}\n", "sep")
                elif stripped.isupper() and len(stripped) < 60 and stripped == stripped.upper() and len(stripped) > 3:
                    self._txt.insert("end", f"\n{stripped}\n", "h2")
                elif stripped.startswith("✗"):
                    self._txt.insert("end", f"  {stripped}\n", "err")
                elif stripped.startswith("✓"):
                    self._txt.insert("end", f"  {stripped}\n", "ok")
                elif stripped.startswith("⚠"):
                    self._txt.insert("end", f"  {stripped}\n", "note")
                elif stripped.startswith("•"):
                    self._txt.insert("end", f"  {stripped}\n", "body")
                elif stripped.startswith("P:"):
                    self._txt.insert("end", f"\n{stripped}\n", "h2")
                elif stripped.startswith("O:"):
                    self._txt.insert("end", f"{stripped}\n", "body")
                elif any(stripped.startswith(p) for p in ("C:\\","KSeF_","/","settings","history","marks")):
                    self._txt.insert("end", f"    {stripped}\n", "mono")
                else:
                    self._txt.insert("end", f"{line}\n", "body")

        self._txt.config(state="disabled")


# ── invoice detail popup ──────────────────────────────────────────────────────
class InvoiceDetailPopup(tk.Toplevel):
    """Wizualne porównanie KSEF / Rejestr / Księga dla wybranej faktury."""

    _SOURCES = [
        ("KSeF",           ["Netto (KSeF)","VAT (KSeF)","Brutto (KSeF)"],            ACCENT),
        ("Rejestr VAT",    ["Netto (VAT rej.)","VAT (VAT rej.)","Brutto (VAT rej.)"], "#3b82f6"),
        ("Rejestr sprz.",  ["Netto (rej. sp.)","VAT (rej. sp.)","Brutto (rej. sp.)"], "#8b5cf6"),
        ("Rejestr",        ["Netto (rejestr)","VAT (rejestr)","Brutto (rejestr)"],    "#3b82f6"),
        ("Księga",         ["Kwota w KSIEGA","% kosztu"],                              "#22c55e"),
    ]
    _HDR_KEYS   = {"Nr faktury","Data wyst.","Data","Sprzedawca","Nabywca",
                   "Kontrahent","NIP","NIP (KSeF)","NIP (rejestr)","Waluta","Nr KSeF","Wystawca"}
    _DELTA_KEYS = {"Δ Netto","Δ VAT","Δ Brutto","Rodzaj błędu","Różnica"}

    def __init__(self, parent, row, check_kind="error", data=None):
        super().__init__(parent)
        self.configure(bg=BG)
        self.resizable(True, True)
        self.minsize(460, 300)
        nr = row.get("Nr faktury","") or row.get("Nr KSeF","faktura")
        self.title(f"Szczegóły — {nr}")
        self._row      = row
        self._kind_col = {"error":ERR,"warning":WARN,"ok":OK}.get(check_kind, TXT2)
        # żywe powiązanie KSIEGA–REJESTR–KSEF (gdy dostępne dane analizy)
        self._link = None
        if data:
            try:
                from checks import lookup_invoice
                self._link = lookup_invoice(
                    data, row.get("Nr faktury") or row.get("NUMER") or row.get("Nr KSeF"))
            except Exception:
                self._link = None
        self._build()
        self.lift(); self.focus_force()

    def _build(self):
        row = self._row
        all_src_keys = {f for _,fs,_ in self._SOURCES for f in fs}
        active  = [(l,fs,c) for l,fs,c in self._SOURCES if any(row.get(f) for f in fs)]
        delta_v = {k:v for k,v in row.items() if k in self._DELTA_KEYS and v}
        extra_v = {k:v for k,v in row.items()
                   if k not in self._HDR_KEYS and k not in self._DELTA_KEYS
                   and k not in all_src_keys and v}

        # ── nagłówek faktury ─────────────────────────────────────────────
        hdr = tk.Frame(self, bg=BG2)
        hdr.pack(fill="x")
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        nr   = row.get("Nr faktury","") or row.get("Nr KSeF","")
        date = row.get("Data wyst.","") or row.get("Data","")
        name = (row.get("Sprzedawca","") or row.get("Nabywca","")
                or row.get("Kontrahent","") or row.get("Wystawca",""))

        top_r = tk.Frame(hdr, bg=BG2)
        top_r.pack(fill="x", padx=14, pady=(10,4))
        if nr:
            tk.Label(top_r, text=nr, font=FMED, bg=BG2, fg=TXT).pack(side="left")
        if date:
            tk.Label(top_r, text=f"  ·  {date}", font=(_SYS,11), bg=BG2, fg=TXT2).pack(side="left")
        tk.Button(top_r, text="✕", font=FSM, bg=BG2, fg=TXT3,
                  activebackground=BG3, relief="flat", padx=8, pady=2,
                  cursor="hand2", command=self.destroy).pack(side="right")

        sub_r = tk.Frame(hdr, bg=BG2)
        sub_r.pack(fill="x", padx=14, pady=(0,10))
        if not name and self._link and self._link.get("ksef"):
            name = self._link["ksef"].get("kontrahent","")
        if name:
            tk.Label(sub_r, text=name, font=FS, bg=BG2, fg=TXT2).pack(side="left")
        badges = [(k, row[k]) for k in ("NIP","NIP (KSeF)","NIP (rejestr)","Waluta","Nr KSeF")
                  if row.get(k)]
        if self._link:
            TYP = {"Zal":"Zaliczkowa","KorZal":"Korekta zaliczkowej",
                   "Roz":"Rozliczeniowa","Vat":"VAT"}
            kf = self._link.get("ksef") or {}
            if kf.get("typ") and not row.get("Typ"):
                badges.append(("Typ", TYP.get(kf["typ"], kf["typ"])))
            if kf.get("kierunek") and not row.get("Kierunek"):
                badges.append(("Kierunek", kf["kierunek"]))
            kn = self._link.get("ksef_nr","")
            if kn and not row.get("Nr KSeF"):
                badges.append(("Nr KSeF", kn if len(kn) <= 30 else kn[:14]+"…"+kn[-13:]))
        for k, v in badges:
            p = tk.Frame(sub_r, bg=BG3)
            p.pack(side="left", padx=(6,0))
            tk.Label(p, text=f"{k}: {v}", font=FSM,
                     bg=BG3, fg=TXT3, padx=6, pady=2).pack()

        # ── karty źródeł danych ──────────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=14, pady=12)

        linked = bool(self._link) and self._build_link_cards(body)
        if active and not linked:
            cf = tk.Frame(body, bg=BG)
            cf.pack(anchor="center")
            for si, (lbl, fields, color) in enumerate(active):
                if si > 0:
                    # strzałka + delta między sąsiednimi kartami
                    arr = tk.Frame(cf, bg=BG)
                    arr.pack(side="left", padx=10, anchor="center")
                    tk.Label(arr, text="⟷", font=(_SYS,16),
                             bg=BG, fg=TXT3).pack(pady=(0,6))
                    if si == 1:
                        # kolejność zgodna z kartami: Netto → VAT → Brutto
                        for dk in ("Δ Netto","Δ VAT","Δ Brutto"):
                            dv = delta_v.get(dk)
                            if dv is None: continue
                            bad = self._nonzero(dv)
                            row = tk.Frame(arr, bg=BG)
                            row.pack(fill="x", pady=2)
                            tk.Label(row, text=dk, font=FSM,
                                     bg=BG, fg=TXT3, width=8, anchor="w").pack(side="left")
                            tk.Label(row, text=dv, font=(_SYS,9,"bold"),
                                     bg=BG, fg=ERR if bad else OK,
                                     anchor="e", width=9).pack(side="left")

                card = tk.Frame(cf, bg=BG2,
                                highlightbackground=color, highlightthickness=2)
                card.pack(side="left", fill="y")
                ch = tk.Frame(card, bg=color)
                ch.pack(fill="x")
                tk.Label(ch, text=lbl, font=(_SYS,10,"bold"),
                         bg=color, fg="#18181b", padx=14, pady=7).pack()
                tk.Frame(card, bg=BG2, height=6).pack()
                for f in fields:
                    val = row.get(f)
                    if not val: continue
                    rf = tk.Frame(card, bg=BG2)
                    rf.pack(fill="x", padx=14, pady=3)
                    dname = (f.replace(" (KSeF)","")
                              .replace(" (VAT rej.)","")
                              .replace(" (rej. sp.)","")
                              .replace("Kwota w KSIEGA","Netto")
                              .replace("SPRZEDAZ_BRUTTO","Brutto"))
                    tk.Label(rf, text=dname, font=FSM, bg=BG2,
                             fg=TXT3, width=9, anchor="e").pack(side="left")
                    tk.Label(rf, text=val, font=(FMONO,11,"bold"),
                             bg=BG2, fg=TXT, padx=8).pack(side="left")
                tk.Frame(card, bg=BG2, height=10).pack()

        # ── rodzaj błędu ─────────────────────────────────────────────────
        if delta_v.get("Rodzaj błędu") and (active or linked):
            rb = tk.Frame(body, bg=BG3)
            rb.pack(fill="x", pady=(10,0))
            tk.Label(rb, text="Rodzaj błędu:", font=FSM, bg=BG3,
                     fg=TXT3, padx=8, pady=4).pack(side="left")
            tk.Label(rb, text=delta_v["Rodzaj błędu"], font=(_SYS,9,"bold"),
                     bg=BG3, fg=ERR, padx=4).pack(side="left")

        # ── pozostałe pola (miesiące, MPP, NIP, daty itd.) ───────────────
        show_extra = dict(extra_v)
        if not active and not linked:
            show_extra.update(delta_v)
        if show_extra:
            ef = tk.Frame(body, bg=BG3)
            ef.pack(fill="x", pady=(10,0))
            for k, v in show_extra.items():
                bad = k.startswith("Δ") and self._nonzero(v)
                rf = tk.Frame(ef, bg=BG3)
                rf.pack(fill="x", padx=2, pady=1)
                tk.Label(rf, text=k, font=FSM, bg=BG3, fg=TXT3,
                         width=24, anchor="e", padx=8, pady=3).pack(side="left")
                tk.Label(rf, text=v, font=(_SYS,9,"bold" if bad else "normal"),
                         bg=BG3, fg=ERR if bad else TXT, padx=4).pack(side="left")

    def _build_link_cards(self, body):
        """Karty KSIEGA – REJESTR – KSeF z żywego odczytu tabel. True gdy narysowano."""
        lk = self._link
        ks, rg, kf = lk.get("ksiega"), lk.get("rejestr"), lk.get("ksef")
        if not (ks or rg or kf):
            return False
        TOL = 0.05

        def money(v):
            try:    return f"{float(v):,.2f}"
            except: return "—"

        ks_rows, rg_rows, kf_rows = [], [], []
        if ks:
            ks_rows = [("Netto", money(ks["netto"]))]
            du = ks.get("data_ujecia") or ks.get("data")
            if du: ks_rows.append(("Ujęcie", du))
        if rg:
            if rg.get("netto") is not None: rg_rows.append(("Netto", money(rg["netto"])))
            if rg.get("vat")   is not None: rg_rows.append(("VAT",   money(rg["vat"])))
            rg_rows.append(("Brutto", money(rg["brutto"])))
            if rg.get("data"): rg_rows.append(("Ujęcie", rg["data"]))
        if kf:
            kf_rows = [("Netto", money(kf["netto"])), ("VAT", money(kf["vat"])),
                       ("Brutto", money(kf["brutto"]))]
            if kf.get("data"): kf_rows.append(("Wyst.", kf["data"]))

        # delty między sąsiednimi kartami (obie strony muszą istnieć)
        gap1, gap2 = [], []
        if ks and rg and rg.get("netto") is not None:
            gap1.append(("Δ Netto", ks["netto"] - rg["netto"]))
        if rg and kf:
            for lbl, key in [("Δ Netto","netto"), ("Δ VAT","vat"), ("Δ Brutto","brutto")]:
                if rg.get(key) is not None:
                    gap2.append((lbl, rg[key] - kf[key]))

        cards = [("KSIEGA",                              "#22c55e", ks, ks_rows,
                  (ks or {}).get("opis","")),
                 ((rg or {}).get("zrodlo") or "REJESTR", "#3b82f6", rg, rg_rows, ""),
                 ("KSeF",                                ACCENT,    kf, kf_rows, "")]

        tk.Label(body, text="Powiązanie  KSIEGA — REJESTR VAT — KSeF",
                 font=FSM, bg=BG, fg=TXT3).pack(anchor="w", pady=(0,6))
        cf = tk.Frame(body, bg=BG)
        cf.pack(anchor="center")

        for i, (lbl, color, src, vrows, footer) in enumerate(cards):
            if i > 0:
                arr = tk.Frame(cf, bg=BG)
                arr.pack(side="left", padx=10, anchor="center")
                tk.Label(arr, text="⟷", font=(_SYS,16), bg=BG, fg=TXT3).pack(pady=(0,6))
                for dk, dv in (gap1 if i == 1 else gap2):
                    bad = abs(dv) > TOL
                    rf = tk.Frame(arr, bg=BG)
                    rf.pack(fill="x", pady=2)
                    tk.Label(rf, text=dk, font=FSM, bg=BG, fg=TXT3,
                             width=8, anchor="w").pack(side="left")
                    tk.Label(rf, text=f"{dv:+,.2f}", font=(_SYS,9,"bold"),
                             bg=BG, fg=ERR if bad else OK,
                             anchor="e", width=10).pack(side="left")

            bcol = color if src else BORDER
            card = tk.Frame(cf, bg=BG2, highlightbackground=bcol, highlightthickness=2)
            card.pack(side="left", fill="y")
            ch = tk.Frame(card, bg=bcol)
            ch.pack(fill="x")
            tk.Label(ch, text=lbl, font=(_SYS,10,"bold"),
                     bg=bcol, fg="#18181b" if src else TXT, padx=14, pady=7).pack()
            tk.Frame(card, bg=BG2, height=6).pack()
            if not src:
                tk.Label(card, text="✗ brak wpisu", font=(_SYS,10,"bold"),
                         bg=BG2, fg=ERR, padx=18, pady=14).pack()
            else:
                for fl, fv in vrows:
                    rf = tk.Frame(card, bg=BG2)
                    rf.pack(fill="x", padx=14, pady=3)
                    tk.Label(rf, text=fl, font=FSM, bg=BG2,
                             fg=TXT3, width=9, anchor="e").pack(side="left")
                    tk.Label(rf, text=fv, font=(FMONO,11,"bold"),
                             bg=BG2, fg=TXT, padx=8).pack(side="left")
                if src.get("rows", 1) > 1:
                    tk.Label(card, text=f"Σ z {src['rows']} wpisów", font=FSM,
                             bg=BG2, fg=TXT3).pack(padx=14, anchor="e")
                if footer:
                    f = footer if len(footer) <= 38 else footer[:37] + "…"
                    tk.Label(card, text=f, font=FSM, bg=BG2,
                             fg=TXT3, padx=10).pack(anchor="w")
            tk.Frame(card, bg=BG2, height=10).pack()
        return True

    @staticmethod
    def _nonzero(v):
        try:    return abs(float(str(v).replace(",","").replace(" ",""))) > 0.005
        except: return bool(v)


# ── main app ──────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("KSeF Checker")
        self.geometry("1320x800")
        self.minsize(1000, 620)
        self.configure(bg=BG)
        self._cfg = cfg_module.load()
        apply_theme(self._cfg.get("theme", "dark"))
        self.configure(bg=BG)
        self._checks = []
        self._selected = None
        self._last_res = None
        self._mdb_path = None
        self._active_cat = None
        self._sort_state = {}
        self._cards_anchor = None
        self._search_text = ""
        self._watch_mtime = None
        # ikona okna
        try:
            self._logo_img256 = _logo_photoimage(_LOGO_256)
            self.iconphoto(True, self._logo_img256)
        except Exception:
            pass
        self._build()
        # sprawdź aktualizacje w tle (max raz dziennie)
        threading.Thread(target=self._check_updates, daemon=True).start()

    # ── aktualizacje ──────────────────────────────────────────────────────
    def _check_updates(self):
        import urllib.request, json, datetime
        today = datetime.date.today().isoformat()
        if self._cfg.get("last_update_check") == today:
            return
        try:
            req = urllib.request.Request(GITHUB_API,
                                         headers={"User-Agent": "KSeF-Checker"})
            with urllib.request.urlopen(req, timeout=5) as r:
                data = json.loads(r.read())
            tag   = data.get("tag_name", "")
            notes = data.get("body", "")[:300]
            # szukaj URL instalatora w assets
            installer_url = None
            for asset in data.get("assets", []):
                if "installer" in asset["name"].lower() and "access" not in asset["name"].lower():
                    installer_url = asset["browser_download_url"]
                    break
            if not installer_url:
                installer_url = data.get("html_url", GITHUB_URL)
            if _ver_tuple(tag) > _ver_tuple(VERSION):
                self.after(0, lambda: self._show_update_banner(
                    tag, notes, installer_url))
            self._cfg["last_update_check"] = today
            cfg_module.save(self._cfg)
        except Exception:
            pass

    def _show_update_banner(self, tag, notes, installer_url):
        banner = tk.Frame(self, bg="#2b1a00", pady=0)
        banner.pack(fill="x", before=self._nb)
        left = tk.Frame(banner, bg="#2b1a00")
        left.pack(side="left", fill="x", expand=True, padx=12, pady=6)
        tk.Label(left, text=f"🔔  Dostępna nowa wersja  {tag}",
                 font=FMED, bg="#2b1a00", fg=WARN).pack(side="left")
        if notes:
            short = notes.split("\n")[0][:80]
            tk.Label(left, text=f"  —  {short}",
                     font=FSM, bg="#2b1a00", fg=TXT2).pack(side="left", padx=6)
        tk.Button(banner, text="  Aktualizuj  ", font=FSM,
                  bg=ACCENT, fg="#18181b",
                  activebackground="#ea6d05", activeforeground="#18181b",
                  relief="flat", padx=10, pady=4,
                  command=lambda: self._start_update(tag, installer_url, banner)).pack(
                  side="right", padx=4, pady=6)
        tk.Button(banner, text="✕", font=FSM,
                  bg="#2b1a00", fg=TXT3,
                  activebackground="#3a2500", activeforeground=TXT,
                  relief="flat", padx=8, pady=4,
                  command=banner.destroy).pack(
                  side="right", padx=(0,4), pady=6)

    def _start_update(self, tag, installer_url, banner=None):
        if banner:
            banner.destroy()
        UpdateDialog(self, self, tag, installer_url)

    # ── layout ────────────────────────────────────────────────────────────
    def _build(self):
        # ── header ────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=BG2, height=68)
        hdr.pack(fill="x"); hdr.pack_propagate(False)

        # separator line na dole headera
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        # logo
        try:
            self._logo_img48 = _logo_photoimage(_LOGO_48)
            tk.Label(hdr, image=self._logo_img48,
                     bg=BG2).pack(side="left", padx=(20,10), pady=10)
        except Exception:
            pass

        # nazwa + wersja
        name_col = tk.Frame(hdr, bg=BG2)
        name_col.pack(side="left", pady=12)
        title_row = tk.Frame(name_col, bg=BG2)
        title_row.pack(anchor="w")
        tk.Label(title_row, text="KSeF Checker", font=FBIG,
                 bg=BG2, fg=TXT).pack(side="left")
        tk.Label(title_row, text=f"  v{VERSION}", font=(_SYS, 9),
                 bg=ACCENT, fg="#18181b", padx=7, pady=2).pack(side="left", padx=8)
        tk.Label(name_col, text="weryfikator spójności bazy MDB",
                 font=(_SYS, 9), bg=BG2, fg=TXT3).pack(anchor="w", pady=(2,0))

        # przyciski po prawej
        def _hdr_btn(text, cmd):
            b = tk.Button(hdr, text=text, font=FSM,
                          bg=BG3, fg=TXT2,
                          activebackground=BG4, activeforeground=TXT,
                          relief="flat", padx=14, pady=7,
                          cursor="hand2", command=cmd)
            b.pack(side="right", padx=4, pady=12)
            b.bind("<Enter>", lambda e: b.config(fg=TXT))
            b.bind("<Leave>", lambda e: b.config(fg=TXT2))
            return b

        _hdr_btn("⚙  Ustawienia", self._open_settings)
        _hdr_btn("🕐  Historia",   self._open_history)
        _hdr_btn("❓  Pomoc",      self._open_help)
        _cur = self._cfg.get("theme","dark")
        _icon = THEMES[_cur]["THEME_ICON"]
        _lbl  = THEMES[_cur]["THEME_LABEL"]
        self._theme_btn = _hdr_btn(f"{_icon}  {_lbl}", self._toggle_theme)
        self._theme_btn.config(fg=ACCENT)

        # ── notebook ──────────────────────────────────────────────────────
        nb_style = ttk.Style(self); nb_style.theme_use("clam")
        nb_style.configure("App.TNotebook", background=BG,
                           borderwidth=0, tabmargins=[0,0,0,0])
        nb_style.configure("App.TNotebook.Tab",
                           background=BG2, foreground=TXT3,
                           font=(_SYS, 10), padding=[20, 8],
                           borderwidth=0)
        nb_style.map("App.TNotebook.Tab",
                     background=[("selected", BG)],
                     foreground=[("selected", TXT)],
                     font=[("selected", (_SYS, 10, "bold"))])
        nb = ttk.Notebook(self, style="App.TNotebook")
        nb.pack(fill="both", expand=True)

        # Tab 1: single file
        self._single_tab = tk.Frame(nb, bg=BG)
        nb.add(self._single_tab, text="  Pojedyncza baza  ")

        # Tab 2: batch
        self._nb = nb
        self._batch_tab = BatchTab(nb, cfg_getter=lambda: self._cfg,
                                   on_open_single=self._open_in_single)
        nb.add(self._batch_tab, text="  Wiele baz (batch)  ")


        # ── build single tab ──────────────────────────────────────────────
        parent = self._single_tab

        # ── toolbar ───────────────────────────────────────────────────────
        tb = tk.Frame(parent, bg=BG2, pady=0)
        tb.pack(fill="x")
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x")

        # inner row z paddingiem
        tb_inner = tk.Frame(tb, bg=BG2)
        tb_inner.pack(fill="x", padx=16, pady=10)

        # file button — wygląda jak pill
        self.file_lbl = tk.Button(tb_inner,
            text="📂  Otwórz plik MDB", font=FS,
            bg=BG3, fg=TXT2,
            activebackground=BG4, activeforeground=TXT,
            relief="flat", padx=14, pady=7,
            cursor="hand2", command=self._pick_file)
        self.file_lbl.pack(side="left")
        self.file_lbl.bind("<Enter>", lambda e: self.file_lbl.config(fg=ACCENT))
        self.file_lbl.bind("<Leave>",
            lambda e: self.file_lbl.config(fg=A2 if self._mdb_path else TXT2))

        # separator
        tk.Frame(tb_inner, bg=BORDER, width=1).pack(side="left", fill="y", padx=14)

        # segmented period control: Cały | Miesiąc | Kwartał
        self.period_var = tk.StringVar(value=self._cfg.get("default_period","month"))
        seg = tk.Frame(tb_inner, bg=BG3)
        seg.pack(side="left")
        self._btn_all = tk.Button(seg, text="Cały", font=FS,
            bg=BG3, fg=TXT2, activebackground=BG4, activeforeground=TXT,
            relief="flat", padx=14, pady=7, cursor="hand2",
            command=lambda: (self.period_var.set("all"), self._toggle_period()))
        self._btn_all.pack(side="left")
        tk.Frame(seg, bg=BORDER, width=1).pack(side="left", fill="y")
        self._btn_month = tk.Button(seg, text="Miesiąc", font=FS,
            bg=BG3, fg=TXT2, activebackground=BG4, activeforeground=TXT,
            relief="flat", padx=14, pady=7, cursor="hand2",
            command=lambda: (self.period_var.set("month"), self._toggle_period()))
        self._btn_month.pack(side="left")
        tk.Frame(seg, bg=BORDER, width=1).pack(side="left", fill="y")
        self._btn_quarter = tk.Button(seg, text="Kwartał", font=FS,
            bg=BG3, fg=TXT2, activebackground=BG4, activeforeground=TXT,
            relief="flat", padx=14, pady=7, cursor="hand2",
            command=lambda: (self.period_var.set("quarter"), self._toggle_period()))
        self._btn_quarter.pack(side="left")

        tk.Frame(tb_inner, bg=BG2, width=8).pack(side="left")

        # selektory okresu
        import datetime; now = datetime.datetime.now()
        self.month_var   = tk.StringVar(value=MONTHS_PL[now.month-1])
        self.year_var    = tk.StringVar(value=str(now.year))
        self.quarter_var = tk.StringVar(value=f"Q{(now.month-1)//3+1}")
        style = ttk.Style()
        style.configure("TB.TCombobox", fieldbackground=BG3, background=BG3,
                        foreground=TXT, selectbackground=BG4, padding=6)
        self.month_cb = ttk.Combobox(tb_inner, textvariable=self.month_var,
                                      values=MONTHS_PL, state="disabled",
                                      width=11, font=FS)
        self.month_cb.pack(side="left", padx=(0,4))
        self.quarter_cb = ttk.Combobox(tb_inner, textvariable=self.quarter_var,
                                        values=["Q1","Q2","Q3","Q4"], state="disabled",
                                        width=5, font=FS)
        self.quarter_cb.pack(side="left", padx=(0,4))
        self.year_sp = tk.Spinbox(tb_inner, from_=2020, to=2035,
                                   textvariable=self.year_var, width=5, font=FS,
                                   bg=BG3, fg=TXT, buttonbackground=BG4,
                                   insertbackground=TXT, relief="flat",
                                   state="disabled")
        self.year_sp.pack(side="left")

        # separator
        tk.Frame(tb_inner, bg=BORDER, width=1).pack(side="left", fill="y", padx=14)

        # action buttons
        self.btn = tk.Button(tb_inner, text="▶  Sprawdź", font=FMED,
            bg=ACCENT, fg="#18181b",
            activebackground="#ea6d05", activeforeground="#18181b",
            relief="flat", padx=18, pady=7,
            cursor="hand2", state="disabled", command=self._run)
        self.btn.pack(side="left", padx=(0,6))

        self.export_btn = tk.Button(tb_inner, text="↓  Excel", font=FB,
            bg=BG3, fg=A2,
            activebackground=BG4, activeforeground=A2,
            relief="flat", padx=14, pady=7,
            cursor="hand2", state="disabled", command=self._export)
        self.export_btn.pack(side="left")

        self.status_lbl = tk.Label(tb_inner, text="", font=FS, bg=BG2, fg=TXT)
        self.status_lbl.pack(side="left", padx=12)

        # ttk global styles
        style = ttk.Style(self); style.theme_use("clam")
        style.configure("P.Horizontal.TProgressbar",
                        troughcolor=BG3, background=ACCENT,
                        bordercolor=BG3, lightcolor=ACCENT, darkcolor=ACCENT)
        style.configure("Treeview",
            background=BG3, foreground=TXT, fieldbackground=BG3,
            bordercolor=BORDER, borderwidth=0, rowheight=28,
            font=(_SYS, 9))
        style.configure("Treeview.Heading",
            background=BG2, foreground=TXT2, relief="flat",
            font=(_SYS, 9, "bold"), padding=[8, 6])
        style.map("Treeview",
            background=[("selected", BG4)],
            foreground=[("selected", ACCENT)])
        self.progress = ttk.Progressbar(tb_inner, style="P.Horizontal.TProgressbar",
                                         mode="indeterminate", length=140)
        self._toggle_period()
        # results area for single tab

        # main split
        main = tk.Frame(parent, bg=BG)
        main.pack(fill="both", expand=True)

        # LEFT scrollable
        left_w = self._cfg.get("left_panel_width", 430)
        left = tk.Frame(main, bg=BG, width=left_w)
        left.pack(side="left", fill="y"); left.pack_propagate(False)

        # ── category filter strip ─────────────────────────────────────────
        self._cat_filter_frame = tk.Frame(left, bg=BG2)
        self._cat_filter_frame.pack(fill="x")
        tk.Frame(left, bg=BORDER, height=1).pack(fill="x")
        self._cat_btns = {}
        _fi  = tk.Frame(self._cat_filter_frame, bg=BG2)
        _fi.pack(fill="x", padx=6, pady=(5,3))
        _fi2 = tk.Frame(self._cat_filter_frame, bg=BG2)
        _fi2.pack(fill="x", padx=6, pady=(0,5))
        _all_btn = tk.Button(_fi, text="Wszystkie", font=FSM,
            bg=ACCENT, fg="#18181b", activebackground="#ea6d05",
            relief="flat", padx=8, pady=3, cursor="hand2",
            command=lambda: self._set_cat_filter(None))
        _all_btn.pack(side="left", padx=(0,3))
        self._cat_btns[None] = _all_btn
        for _cat in _CAT_ORDER[:3]:
            _b = tk.Button(_fi, text=_cat, font=FSM,
                bg=BG3, fg=TXT2, activebackground=BG4, activeforeground=TXT,
                relief="flat", padx=8, pady=3, cursor="hand2",
                command=lambda c=_cat: self._set_cat_filter(c))
            _b.pack(side="left", padx=(0,3))
            self._cat_btns[_cat] = _b
        for _cat in _CAT_ORDER[3:]:
            _b = tk.Button(_fi2, text=_cat, font=FSM,
                bg=BG3, fg=TXT2, activebackground=BG4, activeforeground=TXT,
                relief="flat", padx=8, pady=3, cursor="hand2",
                command=lambda c=_cat: self._set_cat_filter(c))
            _b.pack(side="left", padx=(0,3))
            self._cat_btns[_cat] = _b

        # ── pasek wyszukiwania ────────────────────────────────────────────
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._on_search())
        _sf = tk.Frame(left, bg=BG2)
        _sf.pack(fill="x")
        tk.Frame(left, bg=BORDER, height=1).pack(fill="x")
        _sf_in = tk.Frame(_sf, bg=BG3, padx=4, pady=0)
        _sf_in.pack(fill="x", padx=8, pady=6)
        tk.Label(_sf_in, text="🔍", font=FSM, bg=BG3, fg=TXT3, padx=2).pack(side="left")
        tk.Entry(_sf_in, textvariable=self._search_var, font=FSM,
                 bg=BG3, fg=TXT, insertbackground=TXT,
                 relief="flat", bd=3).pack(side="left", fill="x", expand=True)
        self._search_clear_btn = tk.Button(_sf_in, text="✕", font=(_SYS,8),
            bg=BG3, fg=TXT3, activebackground=BG3, activeforeground=TXT,
            relief="flat", padx=6, cursor="hand2",
            command=lambda: self._search_var.set(""))
        # pokazywany tylko gdy jest tekst (zarządzane przez _on_search)

        # ── stopka sum netto (dół lewego panelu) ─────────────────────────
        tk.Frame(left, bg=BORDER, height=1).pack(side="bottom", fill="x")
        self._netto_footer = tk.Frame(left, bg=BG2)
        self._netto_footer.pack(side="bottom", fill="x")

        left_sb = ttk.Scrollbar(left, orient="vertical")
        left_sb.pack(side="right", fill="y")
        self._left_cv = tk.Canvas(left, bg=BG, highlightthickness=0,
                                   yscrollcommand=left_sb.set)
        self._left_cv.pack(side="left", fill="both", expand=True)
        left_sb.config(command=self._left_cv.yview)

        self.list_frame = tk.Frame(self._left_cv, bg=BG)
        self._lcv_id = self._left_cv.create_window(
            (0,0), window=self.list_frame, anchor="nw")
        self.list_frame.bind("<Configure>",
            lambda e: self._left_cv.configure(scrollregion=self._left_cv.bbox("all")))
        self._left_cv.bind("<Configure>",
            lambda e: self._left_cv.itemconfig(self._lcv_id, width=e.width))
        self._left_cv.bind("<MouseWheel>",
            lambda e: self._left_cv.yview_scroll(-1*(e.delta//120),"units"))

        # divider
        tk.Frame(main, bg=BORDER, width=1).pack(side="left", fill="y")

        # RIGHT
        right = tk.Frame(main, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        self.detail_hdr = tk.Frame(right, bg=BG2, height=56)
        self.detail_hdr.pack(fill="x"); self.detail_hdr.pack_propagate(False)
        tk.Frame(self.detail_hdr, bg=BORDER, width=1).pack(side="left", fill="y")
        self.detail_title = tk.Label(self.detail_hdr,
            text="← Wybierz sprawdzenie",
            font=FMED, bg=BG2, fg=TXT3, padx=20)
        self.detail_title.pack(side="left", pady=16)
        tk.Frame(right, bg=BORDER, height=1).pack(fill="x")

        self.tree_frame = tk.Frame(right, bg=BG)
        self.tree_frame.pack(fill="both", expand=True, padx=10, pady=(8,8))

    def _open_in_single(self, path, month, year):
        """Otwórz bazę z batcha w zakładce Pojedyncza baza i uruchom analizę."""
        # przełącz zakładkę
        self._nb.select(self._single_tab)
        # ustaw plik
        self._mdb_path = path
        name = os.path.basename(path)
        short = name if len(name) <= 44 else name[:42] + "…"
        self.file_lbl.config(text=f"📂  {short}", fg=A2)
        self.btn.config(state="normal")
        self._clear_all()
        # ustaw okres
        if month and year:
            self.period_var.set("month")
            self.month_var.set(MONTHS_PL[month - 1])
            self.year_var.set(str(year))
        else:
            self.period_var.set("all")
        self._toggle_period()
        # odpal analizę
        self._run()

    def _toggle_period(self):
        mode       = self.period_var.get()
        is_month   = mode == "month"
        is_quarter = mode == "quarter"

        # pokaż tylko właściwy selektor
        self.month_cb.pack_forget()
        self.quarter_cb.pack_forget()
        self.year_sp.pack_forget()

        if is_month:
            self.month_cb.config(state="readonly")
            self.month_cb.pack(side="left", padx=(0,4))
            self.year_sp.config(state="normal")
            self.year_sp.pack(side="left")
        elif is_quarter:
            self.quarter_cb.config(state="readonly")
            self.quarter_cb.pack(side="left", padx=(0,4))
            self.year_sp.config(state="normal")
            self.year_sp.pack(side="left")
        else:
            self.month_cb.config(state="disabled")
            self.quarter_cb.config(state="disabled")
            self.year_sp.config(state="disabled")

        for btn, val in [(self._btn_all,"all"),(self._btn_month,"month"),(self._btn_quarter,"quarter")]:
            active = (mode == val)
            btn.config(
                bg=ACCENT if active else BG3,
                fg="#18181b" if active else TXT2,
                activebackground="#ea6d05" if active else BG4)

    def _toggle_theme(self):
        new = "light" if self._cfg.get("theme","dark") == "dark" else "dark"
        self._cfg["theme"] = new
        cfg_module.save(self._cfg)
        apply_theme(new)
        # zachowaj stan
        mdb_path = self._mdb_path
        # przebuduj całe UI
        for w in self.winfo_children():
            w.destroy()
        self.configure(bg=BG)
        self._checks = []; self._selected = None
        self._last_res = None; self._mdb_path = None
        try:
            self._logo_img256 = _logo_photoimage(_LOGO_256)
            self.iconphoto(True, self._logo_img256)
        except Exception:
            pass
        self._build()
        # przywróć plik jeśli był
        if mdb_path:
            self._mdb_path = mdb_path
            name  = os.path.basename(mdb_path)
            short = name if len(name) <= 38 else name[:36] + "…"
            self.file_lbl.config(text=f"📂  {short}", fg=A2, bg=BG3)
            self.btn.config(state="normal")
        threading.Thread(target=self._check_updates, daemon=True).start()

    # ── file watcher ─────────────────────────────────────────────────
    def _start_file_watch(self):
        """Śledź zmiany pliku MDB w tle i zaproponuj odświeżenie."""
        self._stop_file_watch()
        if not self._mdb_path or not os.path.exists(self._mdb_path):
            return
        try:
            self._watch_mtime = os.path.getmtime(self._mdb_path)
        except Exception:
            return
        self._watch_stop_evt = threading.Event()

        def _worker(path, stop_evt, mtime_ref):
            while not stop_evt.wait(4.0):
                try:
                    mt = os.path.getmtime(path)
                    if mt != mtime_ref[0]:
                        mtime_ref[0] = mt
                        stop_evt.set()
                        self.after(0, self._show_refresh_banner)
                        break
                except Exception:
                    break

        mtime_ref = [self._watch_mtime]
        t = threading.Thread(target=_worker,
                             args=(self._mdb_path, self._watch_stop_evt, mtime_ref),
                             daemon=True)
        t.start()

    def _stop_file_watch(self):
        if hasattr(self, "_watch_stop_evt"):
            self._watch_stop_evt.set()

    def _show_refresh_banner(self):
        # nie pokazuj podwójnie
        if hasattr(self, "_refresh_banner"):
            try:
                if self._refresh_banner.winfo_exists():
                    return
            except Exception:
                pass
        banner = tk.Frame(self._single_tab, bg="#152515")
        # wstaw przed pierwszym dzieckiem (toolbar)
        children = self._single_tab.winfo_children()
        banner.pack(fill="x", before=children[0] if children else None)
        self._refresh_banner = banner

        tk.Label(banner, text="🔄  Plik bazy danych został zmieniony",
                 font=FS, bg="#152515", fg=OK, padx=14, pady=6).pack(side="left")
        tk.Button(banner, text="  Odśwież analizę  ", font=FSM,
                  bg=ACCENT, fg="#18181b", activebackground="#ea6d05",
                  relief="flat", padx=10, pady=4, cursor="hand2",
                  command=lambda: (banner.destroy(), self._run())
                  ).pack(side="right", padx=8, pady=6)
        tk.Button(banner, text="✕", font=FSM,
                  bg="#152515", fg=TXT3, activebackground="#1f3a1f",
                  relief="flat", padx=10, pady=4, cursor="hand2",
                  command=lambda: (banner.destroy(), self._start_file_watch())
                  ).pack(side="right", padx=(0,4), pady=6)

    def _open_help(self):
        HelpWindow(self)

    def _open_settings(self):
        SettingsWindow(self, self._cfg, self._on_settings_saved, app_ref=self)

    def _on_settings_saved(self, new_cfg):
        self._cfg = new_cfg

    def _open_history(self):
        HistoryWindow(self, self._load_from_history)

    def _load_from_history(self, entry):
        """Wczytaj wyniki z historii bez ponownej analizy."""
        self._nb.select(self._single_tab)
        self._mdb_path = entry["path"]
        name  = entry["filename"]
        short = name if len(name) <= 44 else name[:42] + "…"
        self.file_lbl.config(text=f"🕐  {short}  (historia)", fg=TXT2, bg=BG3)
        self.btn.config(state="normal")
        self._clear_all()

        # ustaw selektor okresu na podstawie zapisanego wpisu
        period = entry.get("period", "")
        import re
        q_match = re.search(r'Q([1-4])\s+(\d{4})', period)
        if q_match:
            self.period_var.set("quarter")
            self.quarter_var.set(f"Q{q_match.group(1)}")
            self.year_var.set(q_match.group(2))
        else:
            month, year = None, None
            for mi, mname in enumerate(MONTHS_PL, 1):
                if mname in period:
                    month = mi
                    try: year = int(period.strip().split()[-1])
                    except Exception: pass
                    break
            if month and year:
                self.period_var.set("month")
                self.month_var.set(MONTHS_PL[month - 1])
                self.year_var.set(str(year))
            else:
                self.period_var.set("all")
        self._toggle_period()

        self._show_results(entry)

    def _pick_file(self):
        path = filedialog.askopenfilename(
            title="Wybierz plik MDB",
            filetypes=[("Access Database","*.mdb *.MDB *.accdb"),("Wszystkie pliki","*.*")])
        if not path: return
        self._mdb_path = path
        name = os.path.basename(path)
        short = name if len(name)<=38 else name[:36]+"…"
        self.file_lbl.config(text=f"📂  {short}", fg=A2, bg=BG3)
        self.btn.config(state="normal")
        self._clear_all()
        # auto-wybór najnowszego miesiąca z danymi (w tle — odczyt bazy)
        def _detect(p=path):
            per = detect_latest_period(p)
            if per:
                self.after(0, lambda: self._apply_detected_period(p, *per))
        threading.Thread(target=_detect, daemon=True).start()

    def _apply_detected_period(self, path, month, year):
        if self._mdb_path != path:
            return  # użytkownik wybrał już inny plik
        self.month_var.set(MONTHS_PL[month-1])
        self.year_var.set(str(year))
        self.quarter_var.set(f"Q{(month-1)//3+1}")

    def _run(self):
        self.btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self._clear_all()
        self.progress.pack(side="left", padx=6)
        self.progress.start(12)
        self.status_lbl.config(text="Analizuję…", fg=TXT2)

        month, year, quarter = None, None, None
        mode = self.period_var.get()
        try:
            year = int(self.year_var.get())
        except ValueError:
            pass
        if mode == "month":
            try: month = MONTHS_PL.index(self.month_var.get()) + 1
            except (ValueError, IndexError): pass
        elif mode == "quarter":
            try: quarter = int(self.quarter_var.get().replace("Q",""))
            except (ValueError, AttributeError): pass

        cfg = dict(self._cfg)

        def worker():
            try:
                res = analyze_mdb(self._mdb_path, month=month, year=year,
                                   cfg=cfg, progress_cb=lambda m: self._set_status(m),
                                   quarter=quarter)
                self.after(0, lambda: self._show_results(res))
            except Exception as ex:
                msg = str(ex)
                self.after(0, lambda: self._show_error(msg))

        threading.Thread(target=worker, daemon=True).start()

    def _set_status(self, msg):
        self.after(0, lambda: self.status_lbl.config(text=msg))

    def _clear_all(self):
        self._stop_file_watch()
        if hasattr(self, "_netto_footer"):
            for w in self._netto_footer.winfo_children(): w.destroy()
        for w in self.list_frame.winfo_children(): w.destroy()
        self._cards_anchor = None
        self._clear_detail()
        self._checks = []; self._selected = None; self._last_res = None
        self._sort_state = {}

    def _clear_detail(self):
        for w in self.tree_frame.winfo_children(): w.destroy()
        self.detail_title.config(
            text="← Kliknij sprawdzenie aby zobaczyć szczegóły", fg=TXT3)

    def _show_error(self, msg):
        self.progress.stop(); self.progress.pack_forget()
        self.btn.config(state="normal")
        self.status_lbl.config(text="Błąd!", fg=ERR)
        tk.Label(self.list_frame, text=f"✗  {msg}", font=FB, bg=BG, fg=ERR,
                 wraplength=380, justify="left").pack(anchor="w", pady=10)

    # ── results ───────────────────────────────────────────────────────────
    def _show_results(self, res):
        self.progress.stop(); self.progress.pack_forget()
        self.btn.config(state="normal")
        self._last_res = res
        self.export_btn.config(state="normal")
        # zapisz do historii (tylko gdy pochodzi z prawdziwej analizy, nie z historii)
        if self._mdb_path and res.get("status"):
            cfg_module.save_history(self._mdb_path, res)

        s_map = {
            "ok":      ("✓ Spójna",  OK,   "#0d2b1a"),
            "warning": ("⚠ Ostrzeżenia", WARN, "#2b1e06"),
            "error":   ("✗ Błędy",   ERR,  "#2b0d0d"),
        }
        stxt, scol, sbg = s_map.get(res["status"], ("?", TXT2, BG3))
        period = res.get("period","")
        self.status_lbl.config(
            text=f"  {stxt}  {'·' if period else ''}  {period}  " if period else f"  {stxt}  ",
            fg=scol, bg=sbg, padx=2, pady=1)

        # summary strip
        s = res.get("summary",{})
        if s:
            sb = tk.Frame(self.list_frame, bg=BG)
            sb.pack(fill="x", pady=(4,6), padx=4)
            metrics = [
                ("KSeF",     s.get("ksef_total","?"), ACCENT),
                ("Zakupy",   s.get("ksef_zakup","?"),  A2),
                ("Sprzedaż", s.get("ksef_sprz","?"),   TXT2),
                ("Księga",   s.get("ksiega","?"),       TXT2),
                ("VAT zak.", s.get("vatzakupy","?"),    TXT2),
            ]
            for lbl, val, chip_fg in metrics:
                chip = tk.Frame(sb, bg=BG3)
                chip.pack(side="left", padx=(0,6), pady=2)
                tk.Label(chip, text=str(val),
                         font=(FMONO, 15, "bold") if val != "?" else (_SYS,14,"bold"),
                         bg=BG3, fg=chip_fg, padx=14, pady=6).pack()
                tk.Label(chip, text=lbl, font=(_SYS, 8),
                         bg=BG3, fg=TXT3, padx=14, pady=4).pack()

            # VAT chipy usunięte — sumy netto w stopce lewego panelu

        self._checks = res.get("checks", [])
        # reset filter + update button styles
        self._active_cat = None
        for key, btn in self._cat_btns.items():
            btn.config(
                bg=ACCENT if key is None else BG3,
                fg="#18181b" if key is None else TXT2,
                activebackground="#ea6d05" if key is None else BG4)
        # anchor frame marks where cards start (summary stays above)
        self._cards_anchor = tk.Frame(self.list_frame, bg=BG, height=0)
        self._cards_anchor.pack()
        self._rebuild_list(auto_select=True)
        self._update_netto_footer(res.get("summary", {}))
        self._start_file_watch()

    def _update_netto_footer(self, s):
        for w in self._netto_footer.winfo_children():
            w.destroy()
        ksef_zak = s.get("ksef_netto_zak")
        ksef_spr = s.get("ksef_netto_spr")
        reg_zak  = s.get("reg_netto_zak")
        reg_spr  = s.get("reg_netto_spr")
        ks_zak   = s.get("ks_netto_zak")
        ks_spr   = s.get("ks_netto_spr")
        if ksef_zak is None: return

        def _fmt(v):
            try: return f"{float(v):,.2f}"
            except: return "—"

        f = self._netto_footer
        # nagłówek
        hdr = tk.Frame(f, bg=BG2)
        hdr.pack(fill="x", padx=10, pady=(6,2))
        tk.Label(hdr, text="SUMY NETTO", font=(_SYS,7,"bold"),
                 bg=BG2, fg=TXT3).pack(side="left")

        # tabela: label | KSIEGA | REJESTR | KSEF
        tbl = tk.Frame(f, bg=BG2)
        tbl.pack(fill="x", padx=8, pady=(0,6))

        cols_hdr = ["", "Księga", "Rejestr", "KSeF"]
        cols_col = [TXT3, TXT2, TXT2, ACCENT]
        for j, (ch, cc) in enumerate(zip(cols_hdr, cols_col)):
            tk.Label(tbl, text=ch, font=(_SYS,8,"bold"), bg=BG2,
                     fg=cc, width=(12 if j==0 else 13), anchor="e").grid(
                     row=0, column=j, padx=(0,2), sticky="e")

        rows = [
            ("Zakupy",   ks_zak,  reg_zak,  ksef_zak, A2),
            ("Sprzedaż", ks_spr,  reg_spr,  ksef_spr, ERR),
        ]
        for i, (lbl, v_ks, v_reg, v_ksef, row_col) in enumerate(rows, 1):
            tk.Label(tbl, text=lbl, font=FSM, bg=BG2,
                     fg=TXT3, width=12, anchor="w").grid(row=i, column=0, sticky="w")
            for j, val in enumerate([v_ks, v_reg, v_ksef], 1):
                tk.Label(tbl, text=_fmt(val), font=(FMONO,9),
                         bg=BG2, fg=row_col, width=13, anchor="e").grid(
                         row=i, column=j, padx=(0,2), sticky="e")

        # ── WDT / WNT — wewnątrzwspólnotowe, liczone osobno (tylko gdy występują) ──
        wdt, wnt = s.get("wdt_netto") or 0, s.get("wnt_netto") or 0
        try:    has_eu = abs(float(wdt)) > 0.005 or abs(float(wnt)) > 0.005
        except Exception: has_eu = False
        if has_eu:
            tk.Frame(f, bg=BORDER, height=1).pack(fill="x", padx=10)
            eh = tk.Frame(f, bg=BG2); eh.pack(fill="x", padx=10, pady=(4,1))
            tk.Label(eh, text="WEWNĄTRZWSPÓLNOTOWE  ·  osobno (nie wliczone wyżej)",
                     font=(_SYS,7,"bold"), bg=BG2, fg=TXT3).pack(side="left")
            etbl = tk.Frame(f, bg=BG2); etbl.pack(fill="x", padx=8, pady=(0,6))
            for i, (lbl, val, col) in enumerate([("WDT · sprzedaż UE", wdt, ERR),
                                                 ("WNT · zakup UE",    wnt, A2)]):
                tk.Label(etbl, text=lbl, font=FSM, bg=BG2, fg=TXT3,
                         width=20, anchor="w").grid(row=i, column=0, sticky="w")
                tk.Label(etbl, text=_fmt(val), font=(FMONO,9), bg=BG2, fg=col,
                         width=13, anchor="e").grid(row=i, column=1, padx=(0,2), sticky="e")

    def _on_search(self):
        self._search_text = self._search_var.get().strip().lower()
        if self._search_text:
            self._search_clear_btn.pack(side="left")
        else:
            self._search_clear_btn.pack_forget()
        self._rebuild_list(auto_select=False)
        # odśwież prawy panel jeśli coś jest zaznaczone
        if self._selected is not None and self._checks:
            self._show_detail(self._checks[self._selected])

    def _set_cat_filter(self, cat):
        self._active_cat = cat
        for key, btn in self._cat_btns.items():
            active = (key == cat)
            btn.config(
                bg=ACCENT if active else BG3,
                fg="#18181b" if active else TXT2,
                activebackground="#ea6d05" if active else BG4)
        self._rebuild_list(auto_select=True)

    def _rebuild_list(self, auto_select=False):
        # destroy only card widgets — everything after _cards_anchor
        anchor = self._cards_anchor
        found = (anchor is None)
        for w in list(self.list_frame.winfo_children()):
            if found:
                w.destroy()
            if w is anchor:
                found = True
        self._selected = None

        kc = {"ok":(OK,"✓"),"warning":(WARN,"⚠"),"error":(ERR,"✗")}
        _kind_cbg = {"ok": BG2, "warning": CARD_WARN, "error": CARD_ERR}

        def _chk_matches_search(chk):
            if not self._search_text: return True
            for row in chk.get("rows", []):
                if any(self._search_text in str(v).lower() for v in row.values()):
                    return True
            # szukaj też w tytule i detalu
            return (self._search_text in chk.get("title","").lower() or
                    self._search_text in chk.get("detail","").lower())

        from collections import OrderedDict
        grouped: OrderedDict = OrderedDict()
        for i, chk in enumerate(self._checks):
            cat = _CAT_MAP.get(chk.get("id",""), "Inne")
            if self._active_cat is None or cat == self._active_cat:
                if _chk_matches_search(chk):
                    grouped.setdefault(cat, []).append((i, chk))

        def _cat_rank(c): return _CAT_ORDER.index(c) if c in _CAT_ORDER else 999
        sorted_groups = sorted(grouped.items(), key=lambda x: _cat_rank(x[0]))

        for cat_name, items in sorted_groups:
            n_err  = sum(1 for _, c in items if c["kind"]=="error")
            n_warn = sum(1 for _, c in items if c["kind"]=="warning")
            cat_col = ERR if n_err else (WARN if n_warn else TXT3)

            cat_hdr = tk.Frame(self.list_frame, bg=BG)
            cat_hdr.pack(fill="x", padx=4, pady=(14,2))
            row_hdr = tk.Frame(cat_hdr, bg=BG)
            row_hdr.pack(fill="x")
            tk.Label(row_hdr, text=cat_name.upper(),
                     font=(_SYS, 8, "bold"), bg=BG, fg=cat_col).pack(side="left")
            tk.Frame(row_hdr, bg=BORDER, height=1).pack(
                side="left", fill="x", expand=True, padx=8)
            if n_err or n_warn:
                badge_txt = "  ".join(
                    ([f"✗ {n_err}"] if n_err else []) +
                    ([f"⚠ {n_warn}"] if n_warn else []))
                tk.Label(row_hdr, text=badge_txt,
                         font=(_SYS, 8, "bold"), bg=BG, fg=cat_col).pack(side="right")

            for i, chk in items:
                col, icon = kc.get(chk["kind"],(TXT2,"·"))
                all_rows  = chk.get("rows",[])
                n_total   = len(all_rows)
                if self._search_text:
                    n = sum(1 for r in all_rows
                            if any(self._search_text in str(v).lower() for v in r.values()))
                else:
                    n = n_total
                card_bg = _kind_cbg.get(chk["kind"], BG2)

                card = tk.Frame(self.list_frame, bg=BG,
                                highlightbackground=col if chk["kind"]!="ok" else BORDER,
                                highlightthickness=1, cursor="hand2")
                card.pack(fill="x", pady=2, padx=4)

                strip = tk.Frame(card, bg=col if chk["kind"]!="ok" else BG4, width=4)
                strip.pack(side="left", fill="y")
                setattr(self, f"_strip_{i}", strip)

                inner = tk.Frame(card, bg=card_bg)
                inner.pack(side="left", fill="both", expand=True, padx=(12,12), pady=10)

                top = tk.Frame(inner, bg=card_bg)
                top.pack(fill="x")

                tk.Label(top, text=icon, font=(_SYS,11,"bold"),
                         bg=card_bg, fg=col, width=2).pack(side="left")
                tk.Label(top, text=chk["title"], font=(_SYS,11,"bold"),
                         bg=card_bg, fg=TXT if chk["kind"]=="ok" else col,
                         anchor="w").pack(side="left", fill="x", expand=True)

                if n_total:
                    badge_bg = col if chk["kind"]!="ok" else BG4
                    badge_fg = "#18181b" if chk["kind"]!="ok" else TXT3
                    badge_txt = (f"  {n}/{n_total}  " if self._search_text and n != n_total
                                 else f"  {n_total}  ")
                    tk.Label(top, text=badge_txt, font=(FMONO, 9, "bold"),
                             bg=badge_bg, fg=badge_fg).pack(side="right", padx=(6,0))

                if chk.get("detail"):
                    tk.Label(inner, text=chk["detail"], font=(_SYS, 9),
                             bg=card_bg, fg=TXT3, anchor="w").pack(fill="x", pady=(4,0))

                def _hover_on(e, _in=inner, _tp=top, _bg=BG_HOVER):
                    _in.config(bg=_bg); _tp.config(bg=_bg)
                    for w in list(_tp.winfo_children()) + list(_in.winfo_children()):
                        try: w.config(bg=_bg)
                        except: pass

                def _hover_off(e, _in=inner, _tp=top, _bg=card_bg, _c=card):
                    rx = e.widget.winfo_rootx() + e.x - _c.winfo_rootx()
                    ry = e.widget.winfo_rooty() + e.y - _c.winfo_rooty()
                    if not (0 <= rx < _c.winfo_width() and 0 <= ry < _c.winfo_height()):
                        _in.config(bg=_bg); _tp.config(bg=_bg)
                        for w in list(_tp.winfo_children()) + list(_in.winfo_children()):
                            try: w.config(bg=_bg)
                            except: pass

                def on_click(e, idx=i): self._select(idx)

                for w in [card, strip, inner, top] + \
                          list(top.winfo_children()) + list(inner.winfo_children()):
                    w.bind("<Button-1>", on_click)
                    w.bind("<Enter>", _hover_on)
                    w.bind("<Leave>", _hover_off)

                setattr(self, f"_card_{i}", card)

        if auto_select:
            visible_bad = [i for i, c in enumerate(self._checks)
                           if c["kind"] != "ok" and
                           (self._active_cat is None or
                            _CAT_MAP.get(c.get("id",""),"Inne") == self._active_cat)]
            if visible_bad:
                self.after(100, lambda: self._select(visible_bad[0]))

    def _select(self, idx):
        kc = {"ok":OK,"warning":WARN,"error":ERR}
        if self._selected is not None:
            prev = getattr(self, f"_card_{self._selected}", None)
            prev_strip = getattr(self, f"_strip_{self._selected}", None)
            if prev:
                kind = self._checks[self._selected]["kind"]
                prev.config(highlightbackground=kc.get(kind,TXT2) if kind!="ok" else BORDER,
                            highlightthickness=1)
            if prev_strip:
                kind = self._checks[self._selected]["kind"]
                prev_strip.config(bg=kc.get(kind, BG4) if kind!="ok" else BG4)
        self._selected = idx
        card = getattr(self, f"_card_{idx}", None)
        strip = getattr(self, f"_strip_{idx}", None)
        if card: card.config(highlightbackground=ACCENT, highlightthickness=2)
        if strip: strip.config(bg=ACCENT)
        self._show_detail(self._checks[idx])

    def _show_detail(self, chk):
        self._clear_detail()
        kc = {"ok":OK,"warning":WARN,"error":ERR}
        col  = kc.get(chk["kind"], TXT2)
        icon = {"ok":"✓","warning":"⚠","error":"✗"}.get(chk["kind"],"·")
        self.detail_title.config(text=f"{icon}  {chk['title']}", fg=col)

        if chk.get("explanation"):
            exp_box = tk.Frame(self.tree_frame, bg=BG3)
            exp_box.pack(fill="x", padx=8, pady=(4,4))
            tk.Frame(exp_box, bg=ACCENT, width=3).pack(side="left", fill="y")
            tk.Label(exp_box, text=f"  ℹ  {chk['explanation']}",
                     font=(_SYS,9), bg=BG3, fg=TXT2,
                     anchor="w", justify="left", wraplength=760,
                     padx=8, pady=7).pack(side="left", fill="x", expand=True)

        rows_all = chk.get("rows",[])
        if self._search_text:
            rows = [r for r in rows_all
                    if any(self._search_text in str(v).lower() for v in r.values())]
        else:
            rows = rows_all
        if not rows:
            tk.Label(self.tree_frame,
                     text="✓  Brak nieprawidłowości w tej kategorii.",
                     font=FB, bg=BG, fg=OK).pack(anchor="w", pady=20, padx=12)
            return

        cols = list(rows[0].keys())
        chk_id = chk.get("id", id(chk))
        sort_col, sort_asc = self._sort_state.get(chk_id, (None, True))

        # sort rows if column selected
        rows_sorted = list(rows)
        if sort_col and sort_col in cols:
            def _sk(r):
                v = str(r.get(sort_col,"") or "")
                try: return (0, float(v.replace(",","").replace(" ","")))
                except: return (1, v.lower())
            rows_sorted.sort(key=_sk, reverse=not sort_asc)

        # wczytaj istniejące oznaczenia dla tego sprawdzenia
        mdb_key   = self._mdb_path or ""
        cur_marks = cfg_module.get_marks(mdb_key, chk_id)

        def _row_key(row):
            for k in ("Nr faktury","Nr KSeF","KSEF","Numer"):
                if row.get(k): return str(row[k])
            return str(next((v for v in row.values() if v), ""))

        # count + quick-sort buttons
        count_row = tk.Frame(self.tree_frame, bg=BG2)
        count_row.pack(fill="x", padx=8, pady=(4,4))
        tk.Label(count_row, text="Znaleziono:", font=(_SYS,9), bg=BG2, fg=TXT3).pack(side="left")
        _cnt = (f" {len(rows)}/{len(rows_all)} " if self._search_text and len(rows) != len(rows_all)
                else f" {len(rows)} ")
        tk.Label(count_row, text=_cnt, font=("Consolas",9,"bold"),
                 bg=WARN if chk["kind"]=="warning" else (ERR if chk["kind"]=="error" else OK),
                 fg="#18181b").pack(side="left", padx=4)
        tk.Label(count_row, text="wierszy", font=(_SYS,9), bg=BG2, fg=TXT3).pack(side="left")
        if self._search_text:
            tk.Label(count_row, text=f"🔍 \"{self._search_var.get()}\"",
                     font=FSM, bg=BG2, fg=ACCENT).pack(side="left", padx=(8,0))

        # licznik oznaczeń (aktualizowany dynamicznie)
        mark_lbl_var = tk.StringVar()
        mark_lbl = tk.Label(count_row, textvariable=mark_lbl_var, font=FSM,
                            bg=BG2, fg=MARKED_FG)
        mark_lbl.pack(side="left", padx=(10,0))

        def _update_mark_lbl():
            n = len(cfg_module.get_marks(mdb_key, chk_id))
            mark_lbl_var.set(f"✓ {n} zaznaczonych" if n else "")

        # przycisk "Odznacz wszystkie"
        def _clear_all_marks(_chk=chk):
            cfg_module.clear_marks(mdb_key, chk_id)
            self._show_detail(_chk)

        _update_mark_lbl()
        if cur_marks:
            tk.Button(count_row, text="Odznacz wszystkie", font=FSM,
                      bg=BG3, fg=TXT3, activebackground=BG4,
                      relief="flat", padx=8, pady=2, cursor="hand2",
                      command=_clear_all_marks).pack(side="left", padx=(6,0))

        date_cols   = [c for c in cols if "ata" in c]
        amount_cols = [c for c in cols if any(k in c for k in ("Brutto","Netto","Kwota","Δ","VAT"))]
        sort_btns = (date_cols[:1] + amount_cols[:1])[:2]
        if sort_btns:
            tk.Label(count_row, text="Sortuj:", font=(_SYS,9),
                     bg=BG2, fg=TXT3).pack(side="left", padx=(14,4))
            for sc in sort_btns:
                is_active = (sc == sort_col)
                arrow = (" ↑" if sort_asc else " ↓") if is_active else ""
                short = sc.replace(" (KSeF)","").replace(" (VAT rej.)","").replace(" (rej. sp.)","")
                def _sort_click(c=sc, _chk=chk, _id=chk_id):
                    cur_col, cur_asc = self._sort_state.get(_id, (None, True))
                    self._sort_state[_id] = (c, not cur_asc if cur_col == c else True)
                    self._show_detail(_chk)
                tk.Button(count_row, text=f"{short}{arrow}", font=FSM,
                          bg=ACCENT if is_active else BG3,
                          fg="#18181b" if is_active else TXT2,
                          activebackground="#ea6d05" if is_active else BG4,
                          relief="flat", padx=8, pady=2, cursor="hand2",
                          command=_sort_click).pack(side="left", padx=(0,3))

        vsb = ttk.Scrollbar(self.tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(self.tree_frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree = ttk.Treeview(self.tree_frame, columns=cols, show="headings",
                             yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)
        tree.pack(side="left", fill="both", expand=True)

        tree.tag_configure("marked",
                           background=MARKED, foreground=MARKED_FG)

        for c in cols:
            w = max(len(str(c)), max((len(str(r.get(c,"")or"")) for r in rows[:100]),default=0))
            px = min(w*8+20, 320)
            is_sorted = (c == sort_col)
            arrow = (" ↑" if sort_asc else " ↓") if is_sorted else ""
            def _hdr_sort(c=c, _chk=chk, _id=chk_id):
                cur_col, cur_asc = self._sort_state.get(_id, (None, True))
                self._sort_state[_id] = (c, not cur_asc if cur_col == c else True)
                self._show_detail(_chk)
            tree.heading(c, text=f"{c}{arrow}", anchor="w", command=_hdr_sort)
            tree.column(c, width=px, minwidth=50, anchor="w", stretch=False)

        for row in rows_sorted:
            rk  = _row_key(row)
            tag = ("marked",) if rk in cur_marks else ()
            tree.insert("", "end",
                        values=tuple(str(row.get(c,"")or"") for c in cols),
                        tags=tag)

        tree.bind("<MouseWheel>", lambda e: tree.yview_scroll(-1*(e.delta//120),"units"))

        def on_right_click(e, _t=tree, _c=cols, _id=chk_id):
            item = _t.identify_row(e.y)
            if not item: return
            vals = _t.item(item, "values")
            rkey = vals[0] if vals else ""
            is_m = cfg_module.toggle_mark(mdb_key, _id, rkey)
            _t.item(item, tags=("marked",) if is_m else ())
            # wymuś natychmiastowe przerysowanie (Windows TTK bug)
            _t.update_idletasks()
            _t.yview_moveto(_t.yview()[0])
            _update_mark_lbl()
            return "break"

        tree.bind("<Button-3>", on_right_click)

        def on_dbl(e, _t=tree, _c=cols, _chk=chk):
            item = _t.identify_row(e.y)
            if not item: return
            vals = _t.item(item, "values")
            row_dict = {_c[i]: vals[i] for i in range(len(_c)) if i < len(vals)}
            # _data jest tylko po świeżej analizie (nie z historii) — wtedy popup
            # pokazuje żywe powiązanie KSIEGA–REJESTR–KSEF
            data = (self._last_res or {}).get("_data")
            InvoiceDetailPopup(self, row_dict, check_kind=_chk["kind"], data=data)
        tree.bind("<Double-Button-1>", on_dbl)

        tk.Label(self.tree_frame,
                 text="ℹ  Nagłówek = sortuj  |  Dwuklik = szczegóły  |  Prawy klik = zaznacz / odznacz",
                 font=FSM, bg=BG, fg=TXT3, anchor="e").pack(
                 side="bottom", fill="x", padx=8, pady=2)

    # ── export ────────────────────────────────────────────────────────────
    def _export(self):
        if not self._last_res: return
        import os
        default = os.path.splitext(os.path.basename(self._mdb_path))[0]
        period  = self._last_res.get("period","").replace(" ","_")
        path = filedialog.asksaveasfilename(
            title="Zapisz raport Excel",
            defaultextension=".xlsx",
            initialfile=f"KSeF_{default}_{period}.xlsx",
            filetypes=[("Excel","*.xlsx"),("Wszystkie pliki","*.*")])
        if not path: return
        try:
            _export_excel(self._last_res, self._mdb_path, path)
            self.status_lbl.config(
                text=f"✓ Zapisano: {os.path.basename(path)}", fg=OK)
        except Exception as ex:
            import traceback
            self.status_lbl.config(text=f"✗ Błąd eksportu", fg=ERR)
            from tkinter import messagebox
            messagebox.showerror("Błąd eksportu Excel",
                                 f"{ex}\n\n{traceback.format_exc()[-600:]}")


# ── Excel export ──────────────────────────────────────────────────────────────
def _export_excel(res, mdb_path, save_path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import os, datetime

    problems = [c for c in res.get("checks", []) if c["kind"] in ("error", "warning")]

    wb = Workbook()
    wb.remove(wb.active)

    SF  = "Segoe UI"
    def fill(h): return PatternFill("solid", fgColor=h)
    def fnt(h, bold=False, sz=9): return Font(name=SF, color=h, bold=bold, size=sz)
    thin = Side(style="thin", color="dddddd")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ERR_C, WARN_C, HEAD_C = "C0392B", "E67E22", "2C3E50"

    s   = res.get("summary", {})
    now = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")

    # ══════════════════════════════════════════════════════════════════════
    # Arkusz 1 — Podsumowanie problemów
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Raport")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # Nagłówek
    ws.merge_cells("A1:F1")
    ws["A1"] = f"KSeF Checker — {os.path.basename(mdb_path)}   |   {res.get('period','?')}   |   {now}"
    ws["A1"].font = Font(name=SF, color="FFFFFF", bold=True, size=11)
    ws["A1"].fill = fill("1a1a1a")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    # Statystyki w jednej linii
    ws.merge_cells("A2:F2")
    n_err  = sum(1 for c in problems if c["kind"] == "error")
    n_warn = sum(1 for c in problems if c["kind"] == "warning")
    stat = f"Błędy: {n_err}   Ostrzeżenia: {n_warn}   |   KSeF: {s.get('ksef_total','?')}   Zakupy: {s.get('ksef_zakup','?')}   Sprzedaż: {s.get('ksef_sprz','?')}   Księga: {s.get('ksiega','?')}   VAT zak.: {s.get('vatzakupy','?')}"
    ws["A2"] = stat
    ws["A2"].font = Font(name=SF, color="a0a0a0", size=8)
    ws["A2"].fill = fill("2a2a2a")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    if not problems:
        ws.merge_cells("A4:F4")
        ws["A4"] = "✓  Brak błędów i ostrzeżeń — baza spójna."
        ws["A4"].font = Font(name=SF, color="27AE60", bold=True, size=11)
        wb.save(save_path)
        return

    # Nagłówki tabeli
    cols_hdr = ["", "Typ", "Problem", "Liczba rekordów", "Szczegóły", "Wyjaśnienie"]
    col_w    = [4,   10,   52,         16,                48,           52]
    for j, (h, w) in enumerate(zip(cols_hdr, col_w), 1):
        c = ws.cell(4, j, h)
        c.font      = Font(name=SF, color="FFFFFF", bold=True, size=9)
        c.fill      = fill(HEAD_C)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = brd
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[4].height = 18

    for i, chk in enumerate(problems, 1):
        r   = i + 4
        fc  = ERR_C if chk["kind"] == "error" else WARN_C
        ic  = "✗" if chk["kind"] == "error" else "⚠"
        bg  = "FFF5F5" if chk["kind"] == "error" else "FFFBF0"
        n   = len(chk.get("rows", []))

        for j in range(1, 7):
            ws.cell(r, j).fill   = fill(bg)
            ws.cell(r, j).border = brd
            ws.cell(r, j).alignment = Alignment(vertical="center", indent=1, wrap_text=(j in (3,5,6)))

        ws.cell(r, 1, ic).font  = Font(name=SF, color=fc, bold=True, size=11)
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 2, "BŁĄD" if chk["kind"]=="error" else "OSTRZEŻENIE").font = Font(name=SF, color=fc, bold=True, size=8)
        ws.cell(r, 3, chk["title"]).font = Font(name=SF, color="1a1a1a", bold=True, size=9)
        ws.cell(r, 4, n if n else "—").font = Font(name=SF, color=fc, bold=True, size=9)
        ws.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 5, chk.get("detail","")).font = Font(name=SF, color="444444", size=8)
        ws.cell(r, 6, chk.get("explanation","")).font = Font(name=SF, color="666666", size=8)
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:F{4 + len(problems)}"

    # ══════════════════════════════════════════════════════════════════════
    # Arkusz 2 — Szczegóły (wiersze z błędami)
    # ══════════════════════════════════════════════════════════════════════
    chks_with_rows = [c for c in problems if c.get("rows")]
    if chks_with_rows:
        ws2 = wb.create_sheet("Szczegóły")
        ws2.sheet_view.showGridLines = False
        ws2.sheet_properties.pageSetUpPr.fitToPage = True
        ws2.page_setup.fitToWidth = 1

        cur_row = 1
        for chk in chks_with_rows:
            fc  = ERR_C if chk["kind"] == "error" else WARN_C
            ic  = "✗" if chk["kind"] == "error" else "⚠"
            rows = chk.get("rows", [])
            cols = list(rows[0].keys()) if rows else []
            ncols = max(len(cols), 1)

            # sekcja nagłówkowa
            ws2.merge_cells(start_row=cur_row, start_column=1,
                            end_row=cur_row, end_column=ncols)
            c = ws2.cell(cur_row, 1, f"{ic}  {chk['title']}  ({len(rows)} rekordów)")
            c.font = Font(name=SF, color="FFFFFF", bold=True, size=10)
            c.fill = fill(fc.replace("C0392B","8B0000").replace("E67E22","CC5200"))
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws2.row_dimensions[cur_row].height = 20
            cur_row += 1

            if not cols:
                cur_row += 1
                continue

            # nagłówki kolumn
            for j, col in enumerate(cols, 1):
                c = ws2.cell(cur_row, j, col)
                c.font = Font(name=SF, color="FFFFFF", bold=True, size=8)
                c.fill = fill(HEAD_C)
                c.border = brd
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws2.row_dimensions[cur_row].height = 16
            cur_row += 1

            # dane
            for ri, row in enumerate(rows):
                bg = "FFFFFF" if ri % 2 == 0 else "F7F7F7"
                for j, col in enumerate(cols, 1):
                    val = row.get(col, "") or ""
                    try:
                        val = float(val.replace(",","")) \
                            if val.replace(",","").replace(".","").replace("-","").isdigit() else val
                    except Exception:
                        pass
                    c = ws2.cell(cur_row, j, val)
                    c.font   = Font(name=SF, color="1a1a1a", size=8)
                    c.fill   = fill(bg)
                    c.border = brd
                    if isinstance(val, float):
                        c.number_format = "#,##0.00"
                        c.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        c.alignment = Alignment(vertical="center", indent=1)
                ws2.row_dimensions[cur_row].height = 15
                cur_row += 1

            # auto szerokość
            for j, col in enumerate(cols, 1):
                w = max(len(str(col)),
                        max((len(str(r.get(col,"") or "")) for r in rows[:200]), default=0))
                ws2.column_dimensions[get_column_letter(j)].width = min(w + 3, 50)

            cur_row += 1  # odstęp między sekcjami

    wb.save(save_path)


if __name__ == "__main__":
    app = App()
    app.mainloop()
